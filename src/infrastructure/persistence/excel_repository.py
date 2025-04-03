import re
import time
import traceback
import duckdb
import pandas as pd
from typing import Dict, List, Optional, Any, Set, Union
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
import logging
from application.use_cases.inventory.inventory_columns import InventoryColumns
from infrastructure.persistence.cache.limited_cache import LimitedCache
from utils.constant import EXCEL_EXTENSIONS
import os
import json
from functools import lru_cache
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import warnings
import polars as pl

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

logger = logging.getLogger(__name__)

class ExcelRepository:
    """Repository for handling Excel file operations."""
    
    def __init__(
        self, 
        base_path: Optional[Union[str, Path]] = None,
        config_path: Optional[Union[str, Path]] = None
    ):
        """Initialize repository with paths."""
        self.base_path = Path(base_path) if base_path else Path.cwd()
        self.config_path = Path(config_path) if config_path else self.base_path
        self.executor = ThreadPoolExecutor(max_workers=os.cpu_count() or 4)
        self._dataframe_cache = LimitedCache(max_items=20)  # Caché más inteligente
        
        # Inicializar cache para DataFrames
        self._validation_results = {}
        
        self.inventory_required_columns = InventoryColumns.get_required_columns()

        
        # Required columns for audit files (case-insensitive)
        self.audit_required_columns = {
            'FULL PART NUMBER': str,
            'PART#': str,
            'COST TYPE': str,
            'MANUFACTURER': str,
            'MFG NUMBER': str,
            'CONTRACT': str,
            'ORGANIZATION CODE': str,
            'ITEM ORG DESTINATION': str,
            'SERIAL NUMBER CONTROL': str,
            'MFG PART NUM': str,
            'VERTEX PRODUCT CLASS': str,
            'DESCRIPTION': str,
            'CATALOG PRODUCT PART': str,
            'CUSTOMER ID': str,
            'CATEGORY NAME': str,
            'CROSS REFERENCE': str,
            'CROSS REFERENCE DESCRIPTION': str,
            'CROSS REFERENCE TYPE': str,
            'CATEGORY SET NAME': str,
            'CREATED BY': str,
            'CREATION DATE': str,
            'LAST UPDATE DATE': str,
            'LAST UPDATED BY': str,
            'UNSPSC CODE': str,
            'UNSPSC DESCRIPTION': str,
            'ITEM STATUS': str,
        }
        
        # Required columns for inventory files (case-insensitive)
        self.inventory_required_columns = {
            # Columnas clave para identificación
            'ITEM NUMBER': str,                    # Campo principal para coincidencia de partes
            'ORGANIZATION CODE': str,              # Código de organización
            'ORG WAREHOUSE CODE': str,             # Código de almacén
            'SUBINVENTORY CODE': str,              # Código de subinventario
            
            # Columnas de aging - fundamentales para el análisis de inventario
            'AGING 0-30 QUANTITY': float,          # Inventario de 0-30 días
            'AGING 31-60 QUANTITY': float,         # Inventario de 31-60 días
            'AGING 61-90 QUANTITY': float,         # Inventario de 61-90 días
            'AGING 91-120 QUANTITY': float,        # Inventario de 91-120 días
            'AGING 121-150 QUANTITY': float,       # Inventario de 121-150 días
            'AGING 151-180 QUANTITY': float,       # Inventario de 151-180 días
            'AGING 181-365 QUANTITY': float,       # Inventario de 181-365 días
            'AGING OVER 365 QUANTITY': float,      # Inventario mayor a 365 días
            
            # Columnas de valor
            'TOTAL VALUE': float,                  # Valor total del inventario
            'QUANTITY': float,                     # Cantidad total
            
            # Información adicional
            'SERIAL NUMBER': str,                  # Número de serie si aplica
            'ITEM DESCRIPTION': str,               # Descripción del ítem
            'MATERIAL DESIGNATOR': str             # Designador de material (aunque no se use para matching)
        }
        
        self.program_requirements_file = self._resolve_requirements_path()
        
        if not os.environ.get('SKIP_VALIDATION'):
            self._validate_environment()


    def _normalize_column_name(self, col_name: str) -> str:
        return (col_name
                .strip()
                .replace('\n', '')
                .replace('\r', '')
                .replace('\t', '')
                .replace('  ', ' ')
                .upper())

    def _normalize_columns(self, df: Union[pd.DataFrame, pl.DataFrame, pl.LazyFrame]) -> Union[pd.DataFrame, pl.DataFrame, pl.LazyFrame]:
        """
        Normaliza todas las columnas de un DataFrame.
        
        Args:
            df: DataFrame (pandas o polars) a normalizar
            
        Returns:
            DataFrame con columnas normalizadas
        """
        import polars as pl
        
        if isinstance(df, pl.DataFrame):
            return df.rename({
                col: self._normalize_column_name(col) 
                for col in df.columns
            })
        elif isinstance(df, pd.DataFrame):
            df.columns = [self._normalize_column_name(col) for col in df.columns]
            return df
        elif isinstance(df, pl.LazyFrame):
            return df.rename({
                col: self._normalize_column_name(col) 
                for col in df.columns
            })
        else:
            logger.warning(f"Tipo de DataFrame no reconocido: {type(df)}")
            return df
        
    
    def _validate_environment(self) -> None:
        """Validate the environment setup."""
        try:
            if not self.program_requirements_file.exists():
                message = (
                    f"Program requirements file not found at: {self.program_requirements_file}\n"
                    f"Please ensure the file exists in the config directory or set "
                    f"PROGRAM_REQUIREMENTS_PATH environment variable.\n"
                    f"Expected config path: {self.config_path}"
                )
                logger.error(message)
                raise FileNotFoundError(message)
                
            logger.debug("Environment validation successful")
            
        except Exception as e:
            logger.error(f"Environment validation failed: {str(e)}")
            raise

    def _is_inventory_file(self, file_path: Path) -> bool:
        """Check if file is an inventory format."""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            # Verificar primeras 5 filas por si el título está en otra posición
            for row in list(ws.iter_rows(max_row=5)):
                for cell in row:
                    if cell.value and 'WMS' in str(cell.value).upper():
                        wb.close()
                        return True
            wb.close()
            return False
        except Exception as e:
            logger.error(f"Error checking if file is inventory: {str(e)}")
            return False
        
        
    def _read_with_polars(
        self, 
        file_path: Path, 
        is_inventory: bool = False, 
        sheet_name: Optional[str] = None,
        read_mode: str = 'auto'  # 'auto', 'lazy', 'simple'
    ) -> Optional[pd.DataFrame]:
        """
        Método unificado para lectura de archivos Excel con Polars.
        
        Args:
            file_path: Ruta del archivo Excel
            is_inventory: Indica si es un archivo de inventario
            sheet_name: Nombre de la hoja a leer
            read_mode: Modo de lectura ('auto', 'lazy', 'simple')
        
        Returns:
            DataFrame procesado
        """
        start_time = time.time()
        logger.info(f"Iniciando lectura con Polars: {file_path}")
        if isinstance(file_path, str):
            file_path = Path(file_path)
            
        try:
            # Selección de hoja
            if sheet_name is None:
                xl = pd.ExcelFile(file_path)
                sheets = xl.sheet_names
                sheet_name = sheets[0] if sheets else None
                logger.info(f"Usando primera hoja: {sheet_name}")

            # Opciones de lectura comunes
            read_options = {
                "skip_rows": 1,  # Saltar primera fila de título
                "header_row": 1,  # Usar segunda fila como encabezado
            }

            # Determinar modo de lectura
            if read_mode == 'auto':
                file_size_mb = file_path.stat().st_size / (1024 * 1024)
                read_mode = 'lazy' if file_size_mb > 10 or is_inventory else 'simple'

            # Leer con el modo seleccionado
            if read_mode == 'lazy':
                # Procesamiento lazy para archivos grandes
                df_dict = pl.read_excel(file_path, sheet_name=sheet_name, read_options=read_options)
                
                # Seleccionar hoja
                if isinstance(df_dict, dict):
                    sheet_name = list(df_dict.keys())[0]
                    df_pl = df_dict[sheet_name].lazy()
                else:
                    df_pl = df_dict.lazy()

                # Normalizar columnas
                df_pl = df_pl.rename({
                    col: self._normalize_column_name(col) 
                    for col in df_pl.collect_schema().keys()
                })

                # Materializar
                df_pd = df_pl.collect(streaming=True).to_pandas()

            else:
                # Lectura simple para archivos pequeños
                df_pl = pl.read_excel(
                    file_path, 
                    sheet_name=sheet_name, 
                    read_options=read_options
                )
                
                # Convertir a pandas
                df_pd = df_pl.to_pandas()

            # Normalizar columnas
            df_pd.columns = [
                self._normalize_column_name(str(col)) 
                for col in df_pd.columns
            ]

            # Eliminar columnas sin nombre
            df_pd = df_pd.loc[:, ~df_pd.columns.str.contains('^UNNAMED:', case=False, na=False)]

            elapsed = time.time() - start_time
            logger.info(f"Archivo leído con Polars en {elapsed:.2f} segundos. Modo: {read_mode}")
            
            return df_pd

        except Exception as e:
            logger.error(f"Error leyendo archivo con Polars: {str(e)}")
            logger.error(traceback.format_exc())
            return None

    def read_excel_file(
        self, 
        file_path: Path,
        is_inventory: bool = False,
        sheet_name: Optional[str] = None,
        **kwargs
    ) -> pd.DataFrame:
        # Intentar leer con Polars primero
        df = self._read_with_polars(file_path, is_inventory, sheet_name)
        
        # Fallback a Pandas si Polars falla
        if df is None:
            logger.warning("Fallback a lectura con Pandas")
            df = pd.read_excel(
                file_path, 
                skiprows=1, 
                sheet_name=sheet_name, 
                engine='openpyxl',
                dtype=str
            )
            
            # Normalizar columnas
            df.columns = [self._normalize_column_name(str(col)) for col in df.columns]
            df = df.loc[:, ~df.columns.str.contains('^UNNAMED:', case=False, na=False)]
        
        return df
            
    def validate_and_read_file(self, file_path: Union[str, Path], **kwargs) -> pd.DataFrame:
        # Convertir a Path si es string
        file_path = Path(file_path) if isinstance(file_path, str) else file_path
        
        # Generar clave de caché más robusta
        cache_key = (
            str(file_path), 
            kwargs.get('is_inventory', False), 
            kwargs.get('sheet_name')
        )
        
        # Intentar obtener del caché
        cached_df = self._dataframe_cache.get(cache_key)
        if cached_df is not None:
            logger.info(f"Datos recuperados del caché para: {file_path}")
            return cached_df.copy()
        
        # Leer y validar archivo
        df = self.read_excel_file(file_path, **kwargs)
        
        # Validar según tipo de archivo
        if kwargs.get('is_inventory', False):
            self._validate_inventory_columns(df)
        else:
            self._validate_audit_columns(df)
        
        # Guardar en caché
        self._dataframe_cache.set(cache_key, df.copy())
        
        return df
    
    def _validate_dataframe(
        self, 
        df,  # Acepta tanto DataFrame de Pandas como Polars
        is_inventory: bool = False,
        critical_columns: Optional[Set[str]] = None,
        type_checks: Optional[Dict[str, type]] = None
    ) -> Any:  # Devuelve el mismo tipo de DataFrame que recibe
        """
        Validación unificada y flexible de DataFrames compatible con Pandas y Polars.
        
        Args:
            df: DataFrame a validar (Pandas o Polars)
            is_inventory: Indica si es un DataFrame de inventario
            critical_columns: Columnas críticas personalizadas
            type_checks: Validaciones de tipos de datos personalizadas
        
        Raises:
            ValueError: Si la validación falla
        """
        import polars as pl
        
        # 1. Detección del tipo de DataFrame
        is_polars = isinstance(df, (pl.DataFrame, pl.LazyFrame))
        is_pandas = isinstance(df, pd.DataFrame)
        
        if not (is_polars or is_pandas):
            raise ValueError("DataFrame debe ser Pandas o Polars")
        
        # 2. Validación de DataFrame vacío
        if (is_polars and df.is_empty()) or (is_pandas and df.empty):
            raise ValueError("DataFrame está vacío")
        
        # 3. Definir columnas críticas
        if critical_columns is None:
            critical_columns = (
                set(self.inventory_required_columns.keys()) 
                if is_inventory 
                else {'Full Part Number', 'Organization Code', 'Serial Number Control'}
            )
        
        # 4. Función flexible para obtener columnas
        def get_columns(dataframe):
            if is_polars:
                return dataframe.columns
            return list(dataframe.columns)
        
        # 5. Función flexible para verificar existencia de columna
        def column_exists(dataframe, column):
            columns = get_columns(dataframe)
            return any(
                col.lower().replace(' ', '') == column.lower().replace(' ', '') 
                for col in columns
            )
        
        # 6. Verificar columnas críticas
        missing_columns = [
            col for col in critical_columns 
            if not column_exists(df, col)
        ]
        
        if missing_columns:
            raise ValueError(
                f"{'Inventory' if is_inventory else 'Audit'} file missing critical columns: {missing_columns}"
            )
        
        # 7. Conversión y validación de tipos de datos
        def convert_column_type(dataframe, column, expected_type):
            if is_polars:
                # Estrategia para Polars
                return dataframe.with_columns(
                    pl.col(column).cast(pl.Float64 if expected_type == float else pl.Utf8, strict=False)
                )
            else:
                # Estrategia para Pandas
                dataframe[column] = pd.to_numeric(dataframe[column], errors='coerce')
                dataframe[column].fillna(0.0 if expected_type == float else '', inplace=True)
                return dataframe
        
        # 8. Tipos de datos predeterminados
        type_checks = type_checks or {}
        default_type_checks = (
            {col: float for col in get_columns(df) if any(keyword in col.upper() for keyword in ['QUANTITY', 'VALUE', 'AGING'])}
            if is_inventory 
            else {}
        )
        type_checks = {**default_type_checks, **type_checks}
        
        # 9. Aplicar conversión de tipos
        for col, expected_type in type_checks.items():
            matching_column = next(
                (c for c in get_columns(df) if c.lower().replace(' ', '') == col.lower().replace(' ', '')), 
                None
            )
            if matching_column:
                df = convert_column_type(df, matching_column, expected_type)
        
        # 10. Logging de información
        logger.info(f"Validación completada para {'inventario' if is_inventory else 'auditoría'}")
        logger.info(f"Total de registros: {len(df)}")
        logger.info(f"Columnas validadas: {get_columns(df)}")
        
        return df

    def _validate_file_basics(self, file_path: Union[str, Path]) -> Path:
        """Validate basic file requirements."""
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
            
        if path.suffix.lower() not in EXCEL_EXTENSIONS:
            raise ValueError(f"Invalid file format. Must be one of: {EXCEL_EXTENSIONS}")
            
        return path
    
    def validate_input_file(self, file_path: Union[str, Path]) -> bool:
        """Validate audit file format and content."""
        df = self.read_excel_file(file_path, is_inventory=False)
        self._validate_dataframe(df, is_inventory=False)
        return True

    def validate_inventory_file(self, file_path: Union[str, Path]) -> bool:
        """Validate inventory file format and content."""
        df = self.read_excel_file(file_path, is_inventory=True)
        self._validate_dataframe(df, is_inventory=True)
        return True
            
    def _validate_audit_columns(self, df: pd.DataFrame) -> None:
        """Validate required columns for audit files."""
        critical_columns = {'FULL PART NUMBER', 'ORGANIZATION CODE', 'SERIAL NUMBER CONTROL'}
        missing_critical = critical_columns - set(df.columns)
        if missing_critical:
            raise ValueError(f"Audit file missing critical columns: {missing_critical}")
        # Opcional: Loguear columnas faltantes no críticas
        all_missing = set(self.audit_required_columns.keys()) - set(df.columns)
        if all_missing - missing_critical:
            logger.warning(f"Columnas no críticas faltantes en archivo de auditoría: {all_missing - missing_critical}")
            
    def _validate_inventory_columns(self, df: pd.DataFrame) -> None:
        """Valida las columnas requeridas para archivos de inventario."""
        logger.info("=== VALIDACIÓN DE COLUMNAS DE INVENTARIO ===")
        
        # Mostrar columnas disponibles
        logger.info("Columnas disponibles:")
        for col in df.columns:
            logger.info(f"  - {col}")
        
        # Normalizar columnas requeridas
        required_columns = {self._normalize_column_name(col) for col in self.inventory_required_columns.keys()}
        df_columns = {self._normalize_column_name(col) for col in df.columns}
        
        logger.info("\nColumnas requeridas:")
        for col in sorted(required_columns):
            logger.info(f"  - {col}")
        
        # Encontrar columnas faltantes
        missing_columns = required_columns - df_columns
        
        logger.info("\nColumnas faltantes:")
        for col in missing_columns:
            logger.info(f"  - {col}")
        
        if missing_columns:
            raise ValueError(f"Inventory file missing critical columns: {missing_columns}")
        
    def save_excel_file(self, df: pd.DataFrame, file_path: Path) -> None:
        """
        Save DataFrame to Excel with consistent formatting.
        
        Args:
            df: DataFrame to save
            file_path: Path where to save the file
        """
        try:
            # Configure default styles for new workbook
            options = {
                'engine': 'openpyxl',
                'index': False,
                # Remover encoding ya que no es soportado por pandas to_excel
            }
            
            # Save with configured options
            df.to_excel(file_path, **options)
            
            # Apply styles
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            font = Font(name='Calibri', size=11)
            alignment = Alignment(horizontal='general', vertical='bottom')
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = font
                    cell.alignment = alignment
            
            wb.save(file_path)
            
        except Exception as e:
            logger.error(f"Error saving Excel file {file_path}: {str(e)}")
            raise

    def get_program_requirements(self, contract: str) -> Dict[str, Any]:
        try:
            if not self.program_requirements_file.exists():
                raise FileNotFoundError(f"Program requirements file not found: {self.program_requirements_file}")
            
            # Define dtypes to force string reading for org columns
            dtypes = {
                'CONTRACT': str,
                'ORG FOR SERIAL CONTROL COMPARISION': str,
                'ORGANIZATION CODE (PHYSICAL ORGS)': str,
                'ORGANIZATION CODE (DROPSHIP ORGS)': str,
                'ITEM ORG DESTINATION THAT CAN BE USED + OTHER ORGS': str,
                'DOES PROGRAM TRANSACT INTERNATIONALLY?': str,
                'STATUS': str
            }
            
            # Read Excel file with specified dtypes
            df = pd.read_excel(self.program_requirements_file, engine='openpyxl', dtype=dtypes)
            df = self._normalize_columns(df)
            
            # Required columns validation
            required_columns = {
                'CONTRACT': 'Contract ID',
                'ORG FOR SERIAL CONTROL COMPARISION': 'Base Organization',
                'ORGANIZATION CODE (PHYSICAL ORGS)': 'Physical Organizations',
                'ORGANIZATION CODE (DROPSHIP ORGS)': 'Dropship Organizations',
                'ITEM ORG DESTINATION THAT CAN BE USED + OTHER ORGS': 'Item Org Destination',
                'DOES PROGRAM TRANSACT INTERNATIONALLY?': 'International Flag',
                'STATUS': 'Status'
            }
            
            missing_columns = set(required_columns.keys()) - set(df.columns)
            if missing_columns:
                raise ValueError(f"Missing required columns in requirements file: {', '.join(missing_columns)}")
            
            # Contract validation and data retrieval
            contract = str(contract).strip().upper()
            contract_mask = df['CONTRACT'].str.upper() == contract
            program_data = df[contract_mask]
            
            if program_data.empty:
                raise ValueError(f"No requirements found for contract: {contract}")
                
            # Extract base org with validation    
            base_org = program_data['ORG FOR SERIAL CONTROL COMPARISION'].iloc[0]
            if pd.isna(base_org):
                raise ValueError(f"Base organization not specified for contract {contract}")
            
            base_org = str(base_org).strip()
            if base_org.endswith('.0'):
                base_org = base_org[:-2]
            base_org = base_org.zfill(2)
            
            # Process organizations
            destination_orgs_raw = program_data['ITEM ORG DESTINATION THAT CAN BE USED + OTHER ORGS'].iloc[0]
            physical_orgs_raw = program_data['ORGANIZATION CODE (PHYSICAL ORGS)'].iloc[0]
            dropship_orgs_raw = program_data['ORGANIZATION CODE (DROPSHIP ORGS)'].iloc[0]
            
            # Add debug log for raw input
            logger.debug(f"Raw input before processing - Destination: {destination_orgs_raw}")
            logger.debug(f"Raw input before processing - Physical: {physical_orgs_raw}")
            logger.debug(f"Raw input before processing - Dropship: {dropship_orgs_raw}")
            
            destination_orgs = self._process_org_codes(destination_orgs_raw)
            physical_orgs = self._process_org_codes(physical_orgs_raw)
            dropship_orgs = self._process_org_codes(dropship_orgs_raw)
            
            all_orgs = destination_orgs + physical_orgs + dropship_orgs
            unique_orgs = sorted(list(set(all_orgs)))
            
            logger.debug(f"Raw destination orgs: {destination_orgs_raw}")
            logger.debug(f"Processed destination orgs: {destination_orgs}")
            logger.debug(f"Raw physical orgs: {physical_orgs_raw}")
            logger.debug(f"Processed physical orgs: {physical_orgs}")
            logger.debug(f"Raw dropship orgs: {dropship_orgs_raw}")
            logger.debug(f"Processed dropship orgs: {dropship_orgs}")
            logger.debug(f"Combined unique orgs: {unique_orgs}")
            
            if not destination_orgs:
                raise ValueError(f"Item Org Destination organizations are required for contract {contract}")
            
            requirements = {
                'contract': contract,
                'base_org': base_org,
                'org_destination': unique_orgs,
                'physical_orgs': physical_orgs,
                'dropship_orgs': dropship_orgs,
                'international': bool(program_data['DOES PROGRAM TRANSACT INTERNATIONALLY?'].iloc[0]),
                'status': program_data['STATUS'].iloc[0]
            }
            
            logger.debug(f"Program requirements loaded for contract {contract}:")
            logger.debug(f"Base org: {requirements['base_org']}")
            logger.debug(f"Org destination: {requirements['org_destination']}")
            logger.debug(f"Physical orgs: {requirements['physical_orgs']}")
            logger.debug(f"Dropship orgs: {requirements['dropship_orgs']}")
            
            return requirements
        
        except Exception as e:
            logger.error(f"Error retrieving program requirements for contract {contract}: {str(e)}")
            logger.error(f"Stack trace: {traceback.format_exc()}")
            raise

    def _process_org_codes(self, org_codes: str) -> List[str]:
        """
        Procesa códigos de organización con soporte completo para diferentes patrones, prefijos y variaciones.

        Args:
            org_codes: Cadena que contiene códigos de organización

        Returns:
            List[str]: Lista de códigos de organización validados
        """
        try:
            if pd.isna(org_codes):
                return []
                        
            # Convertir a cadena y limpiar valores flotantes
            org_str = str(org_codes).strip()
            if org_str.endswith('.0'):
                org_str = org_str[:-2]
                        
            if not org_str:
                return []
            
            normalized_orgs = set()

            # Separar usando palabras clave comunes como "and", "add", "&", "+"
            add_parts = re.split(r'\s+and\s+add\s+|\s+and\s+|\s*add\s*|\s*&\s*|\s*\+\s*', org_str, flags=re.IGNORECASE)
            
            # Procesar la parte principal (antes de cualquier "and add" o similar)
            main_part = add_parts[0]
            
            # Extraer números de cualquier prefijo, priorizando contexto de "org"
            org_context = re.search(r'(?:org\s+|in\s+org\s+)(\d+(?:[,-]\d+)*)', main_part, re.IGNORECASE)
            if org_context:
                numbers = re.findall(r'\d+', org_context.group(1))  # Ej: "04,132,40" -> ["04", "132", "40"]
            else:
                numbers = re.findall(r'\d+', main_part)
            for num in numbers:
                normalized_orgs.add(num.zfill(2))

            # Manejar casos con "No ORGs" o "per Shannon"
            if "No ORGs" in org_str:
                no_orgs_numbers = re.findall(r'\d+', org_str)
                for num in no_orgs_numbers:
                    normalized_orgs.add(num.zfill(2))

            if "per shannon" in org_str.lower():
                shannon_numbers = re.findall(r'\d+', org_str)
                for num in shannon_numbers:
                    normalized_orgs.add(num.zfill(2))

            # Procesar los números adicionales después de "and add", "and", "add", "&", "+"
            for part in add_parts[1:]:
                logger.debug(f"Procesando parte adicional: '{part}'")
                org_context = re.search(r'(?:org\s+|in\s+org\s+)(\d+(?:[,-]\d+)*)', part, re.IGNORECASE)
                if org_context:
                    additional_numbers = re.findall(r'\d+', org_context.group(1))
                else:
                    additional_numbers = re.findall(r'\d+', part)
                for num in additional_numbers:
                    # Evitar agregar números que sean combinación de códigos previos
                    if not any(num in existing or existing in num for existing in normalized_orgs):
                        normalized_orgs.add(num.zfill(2))
                        logger.debug(f"Agregada organización adicional: {num.zfill(2)}")

            # Retornar resultados ordenados y únicos
            valid_orgs = sorted(normalized_orgs)
            logger.debug(f"Códigos de organización procesados: Entrada='{org_codes}' Salida={valid_orgs}")
            return valid_orgs
                    
        except Exception as e:
            logger.error(f"Error al procesar códigos de organización: {str(e)}")
            logger.error(f"Valor de entrada: {org_codes}")
            return []
            
    @staticmethod
    def extract_organizations(org_str: str) -> List[str]:
        """
        Extraer organizaciones de manera extremadamente robusta.
        Maneja múltiples formatos y variaciones.
        
        Args:
            org_str: String conteniendo códigos de organización
            
        Returns:
            List[str]: Lista ordenada de códigos de organización únicos
        """
        if pd.isna(org_str):
            return []
        
        # Convertir a string y normalizar
        org_str = str(org_str).lower()
        
        # Patrones de extracción extendidos
        extraction_patterns = [
            r'wwt_orgs_(\d+(?:_\d+)*)',
            r'wwt_sing_orgs_(\d+(?:_\d+)*)',
            r'wwt_uk_orgs_(\d+(?:_\d+)*)',
            r'wwt_ind_orgs_(\d+(?:_\d+)*)',
            r'wwt_br_orgs_(\d+(?:_\d+)*)',
            r'wwt_vn_orgs_(\d+(?:_\d+)*)',
            r'telco_(\d+(?:_\d+)*)',
            r'orgs(?:_|\s*)(\d+(?:[,_\s]\d+)*)',
            # Patrones adicionales para casos comunes
            r'org[s]?\s*[:=-]?\s*(\d+(?:[,_\s]\d+)*)',
            r'location[s]?\s*[:=-]?\s*(\d+(?:[,_\s]\d+)*)'
        ]
        
        additional_patterns = [
            r'and\s*add\s*(\d+(?:[,_\s]\d+)*)',
            r'\+\s*(\d+(?:[,_\s]\d+)*)',
            r'with\s*(\d+(?:[,_\s]\d+)*)',
            r'includes?\s*(\d+(?:[,_\s]\d+)*)'
        ]
        
        organizations = set()
        
        # Intentar extraer con todos los patrones
        all_patterns = extraction_patterns + additional_patterns
        for pattern in all_patterns:
            matches = re.findall(pattern, org_str)
            for match in matches:
                orgs = re.findall(r'\d+', str(match))
                organizations.update(orgs)
        
        # Extracción general si no se encontraron organizaciones
        if not organizations:
            direct_orgs = re.findall(r'\b(\d{2,3})\b', org_str)
            organizations.update(direct_orgs)
        
        # Filtrar y validar
        valid_orgs = sorted(set(
            org for org in organizations 
            if len(org) >= 2 and len(org) <= 3
        ))
        
        return valid_orgs
    
    def _resolve_requirements_path(self) -> Path:
        """
        Find the program requirements file path.
        
        Returns:
            Path: Ruta validada al archivo de requerimientos
            
        Raises:
            FileNotFoundError: Si no se encuentra el archivo en ninguna ubicación
        """
        possible_filenames = [
            "Program Requirements - PDM - NPI for Audit.xlsx",
            "Program Requirements - PDM - NPI  for Audit.xlsx"
        ]
        
        logger.debug("Searching for requirements file in possible locations...")
        
        # Rutas posibles con prioridad
        search_paths = [
            self.config_path,
            self.base_path / "config",
            self.base_path,
            Path(__file__).parent.parent.parent / "config"
        ]
        
        # Buscar en todas las combinaciones posibles
        for filename in possible_filenames:
            for base_path in search_paths:
                path = base_path / filename
                logger.debug(f"Checking path: {path}")
                if path.exists():
                    logger.info(f"Found requirements file at: {path}")
                    return path
        
        # Verificar variable de entorno
        env_path = os.environ.get('PROGRAM_REQUIREMENTS_PATH')
        if env_path:
            path = Path(env_path)
            if path.exists():
                logger.info(f"Using requirements file from environment: {path}")
                return path
            else:
                logger.warning(f"Environment path does not exist: {env_path}")
        
        # Usar ruta por defecto como último recurso
        default_path = self.config_path / "Program Requirements - PDM - NPI for Audit.xlsx"
        logger.warning(f"No requirements file found, defaulting to: {default_path}")
        
        if not default_path.exists():
            error_msg = (
                f"Requirements file not found in any location. "
                f"Searched in: {[str(p) for p in search_paths]}"
            )
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        return default_path

    def cleanup(self) -> None:
        """Clean up repository resources."""
        try:
            if hasattr(self, 'executor'):
                logger.debug("Shutting down executor...")
                self.executor.shutdown(wait=True)
        except Exception as e:
            logger.error(f"Error during cleanup: {str(e)}")

    def __del__(self):
        """Ensure cleanup on destruction."""
        self.cleanup()
        
    def read_inventory_file(self, file_path: Path, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Lee y normaliza archivo de inventario explícitamente.
        """
        try:
            logger.info(f"Reading inventory file: {file_path}")
            
            # Usar el método general con flag explícito de inventario
            return self.read_excel_file(
                file_path, 
                is_inventory=True,  # CLAVE: Forzar modo inventario
                sheet_name=sheet_name
            )
                
        except Exception as e:
            logger.error(f"Error reading inventory file: {str(e)}")
            raise
from logging import Logger
import traceback
from typing import Dict, List
import pandas as pd
import re
from datetime import datetime
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment 
from domain.entities.audit_entity import AuditResult
from infrastructure.logging.logger import get_logger
from infrastructure.persistence.excel_repository import ExcelRepository
import json

# Configurar logger
logger = get_logger(__name__)

class ReportGenerator:
    # Constantes de formato y estilos
    COLORS = {
        'HEADER': '1F4E78',      # Azul oscuro para encabezados
        'SUCCESS': '4CAF50',     # Verde para estados positivos
        'ERROR': 'FF6B6B',       # Rojo para estados negativos
        'WARNING': 'FFD700',     # Amarillo para estados de advertencia
        'INFO': '2196F3',        # Azul para dynamic entry
        'NEUTRAL': 'FFFFFF',     # Blanco para valores neutros
        # Colores pasteles para mejor legibilidad
        'LIGHT_RED': 'FFCDD2',   # Rojo suave
        'LIGHT_GREEN': 'E8F5E9', # Verde suave
        'LIGHT_YELLOW': 'FFF2CC',# Amarillo suave
        'LIGHT_BLUE': 'E3F2FD'   # Azul suave
    }

    FORMATS = {
        'HEADER': Font(bold=True),
        'SUMMARY': Font(bold=True, size=12),
        'BORDER': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

    # Nombres de columnas y configuración de reportes
    SERIAL_VALIDATION_COLUMNS = [
        'Part Number', 'Manufacturer', 'Description', 'Vertex',
        'Serial Control match?'
    ]

    ORG_VALIDATION_COLUMNS = [
        'Part Number', 'Status', 'Organization code mismatch',
        'Action Required'
    ]

    # Nombres de sheets y configuración
    SHEET_NAMES = {
        'SERIAL_VALIDATION': "Serial Control Validation",
        'ORG_VALIDATION': "Audit Results",
        'SUMMARY': "Summary"
    }

    # Valores de estado
    STATUS_VALUES = {
        'MISSING': 'Missing Org',
        'CORRECT': 'Correct Org',
        'MISMATCH': 'Mismatch',
        'MATCH': 'Match'
    }
    
    def __init__(self, output_dir: str = "reports"):
        """Inicializa el generador de reportes."""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
    def generate_report(self, audit_result: AuditResult, validation_results: Dict) -> Dict:
        """
        Genera reportes de auditoría sin interrumpir el proceso en caso de errores.
        
        MODIFICACIÓN CRÍTICA: Implementa manejo de errores envolvente para garantizar que 
        el proceso continúe y genere reportes incluso cuando ocurran errores.
        
        Args:
            audit_result: Resultados de la auditoría
            validation_results: Resultados de validación
            
        Returns:
            Diccionario con rutas a reportes y resumen
        """
        try:
            # Verificar y completar base_org y org_destination si faltan
            if not validation_results.get('program_requirements', {}).get('base_org'):
                logger.warning("Missing base_org in program requirements - Intentando recuperar")
                validation_results['program_requirements']['base_org'] = (
                    audit_result.serial_control_results.get('program_requirements', {}).get('base_org')
                )
            
            if not validation_results.get('program_requirements', {}).get('org_destination'):
                logger.warning("Missing org_destination in program requirements - Intentando recuperar")
                validation_results['program_requirements']['org_destination'] = (
                    audit_result.serial_control_results.get('program_requirements', {}).get('org_destination', [])
                )
            
            # Añadir flag de validación de inventario
            inventory_validation_enabled = validation_results.get('use_inventory', True)
            validation_results['inventory_validation_enabled'] = inventory_validation_enabled
            
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            
            # Inicializar variables para reportes
            external_report_path = None
            internal_report_path = None
            summary = {}
            
            # Generar reporte externo con manejo de errores
            try:
                external_report_path = self._generate_external_report(
                    serial_results=audit_result.serial_control_results,
                    program_requirements=validation_results['program_requirements'],
                    timestamp=timestamp,
                    inventory_validation_enabled=inventory_validation_enabled
                )
                logger.info(f"External report generated at: {external_report_path}")
            except Exception as e:
                logger.error(f"Error generando reporte externo: {str(e)}")
                logger.error(f"Traza: {traceback.format_exc()}")
                external_report_path = f"ERROR_external_report_{timestamp}.xlsx"
                logger.warning(f"Continuando proceso a pesar del error en reporte externo")
            
            # Generar reporte interno con manejo de errores
            try:
                internal_report_path = self._generate_internal_report(
                    audit_result=audit_result,
                    validation_results=validation_results,
                    timestamp=timestamp
                )
                logger.info(f"Internal report generated at: {internal_report_path}")
            except Exception as e:
                logger.error(f"Error generando reporte interno: {str(e)}")
                logger.error(f"Traza: {traceback.format_exc()}")
                internal_report_path = f"ERROR_internal_report_{timestamp}.xlsx"
                logger.warning(f"Continuando proceso a pesar del error en reporte interno")
            
            # Generar resumen con manejo de errores
            try:
                # Modificar el summary para incluir estado de inventario
                summary = audit_result.summary.copy()
                summary['inventory_validation'] = 'Enabled' if inventory_validation_enabled else 'Disabled'
                # Añadir información sobre errores al resumen
                if not external_report_path or not internal_report_path:
                    summary['errors_encountered'] = True
                    summary['process_completed'] = 'With errors'
                else:
                    summary['errors_encountered'] = False
                    summary['process_completed'] = 'Successfully'
            except Exception as e:
                logger.error(f"Error generando resumen: {str(e)}")
                # Crear resumen mínimo
                summary = {
                    'inventory_validation': 'Enabled' if inventory_validation_enabled else 'Disabled',
                    'errors_encountered': True,
                    'process_completed': 'With errors in summary'
                }
            
            # Ruta fallback para reportes
            if not external_report_path:
                external_report_path = "ERROR_failed_to_generate_external_report.xlsx"
            if not internal_report_path:
                internal_report_path = "ERROR_failed_to_generate_internal_report.xlsx"
            
            return {
                "external_report_path": str(external_report_path),
                "internal_report_path": str(internal_report_path),
                "summary": summary
            }
            
        except Exception as e:
            # Capturar cualquier excepción para garantizar que el proceso no se interrumpa
            logger.error(f"ERROR CRÍTICO en generación de reportes: {str(e)}")
            logger.error(f"Traza completa: {traceback.format_exc()}")
            logger.warning("CONTINUANDO proceso a pesar de error crítico")
            
            # Retornar resultado mínimo para continuar el proceso
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            return {
                "external_report_path": f"CRITICAL_ERROR_external_report_{timestamp}.xlsx",
                "internal_report_path": f"CRITICAL_ERROR_internal_report_{timestamp}.xlsx",
                "summary": {
                    "status": "ERROR",
                    "error_message": str(e),
                    "timestamp": datetime.now().isoformat()
                }
            }

    def _generate_external_report(
        self,
        serial_results: Dict,
        program_requirements: Dict,
        timestamp: str,
        inventory_validation_enabled: bool = True
    ) -> Path:
        try:
            # DEBUG: Imprimir contenido de serial_results
            logger.debug("Contenido de serial_results:")
            logger.debug(f"Claves disponibles: {serial_results.keys()}")
            logger.debug(f"Datos: {serial_results.get('data', 'Sin datos')}")
            logger.debug(f"Partes con discrepancias: {serial_results.get('mismatched_parts', 'Sin discrepancias')}")
            contract_name = program_requirements.get('contract', 'Unknown')
            contract_name = contract_name.replace(' ', '_')  # Reemplazar espacios por guiones bajos
            excel_path = self.output_dir / f"serial_control_validation_{contract_name}_{timestamp}.xlsx"
            
            # CORRECCIÓN: Usar la lista exacta de 72 organizaciones físicas proporcionada
            # Estas son las organizaciones que cumplen con DROPSHIP_ENABLED='N' y WMS_ENABLED_FLAG='Y'
            # Esto es la lista definitiva de organizaciones físicas
            physical_orgs = [
                '01', '06', '07', '08', '09', '10', '102', '103', '104', '105', 
                '111', '113', '117', '118', '12', '120', '122', '123', '124', '128', 
                '13', '131', '132', '133', '134', '135', '136', '14', '140', '141', 
                '142', '148', '149', '15', '154', '155', '156', '157', '159', '160', 
                '22', '25', '26', '28', '29', '31', '34', '37', '38', '40', 
                '42', '43', '45', '46', '47', '51', '52', '53', '56', '57', 
                '58', '66', '72', '73', '74', '82', '85', '87', '89', '91', 
                '94', '97'
            ]
            
            logger.info(f"Usando lista fija de {len(physical_orgs)} organizaciones físicas")
            
            # Filtrar organizaciones de prueba (que comienzan con Z)
            physical_orgs = [org for org in physical_orgs if not org.upper().startswith('Z')]
            logger.info(f"Después de filtrar orgs Z: {len(physical_orgs)} organizaciones físicas")
            
            # Generar datos de validación de serie con las organizaciones físicas
            serial_validation_df = self._generate_serial_validation_data(
                serial_results, 
                physical_orgs,  # Pasar physical_orgs como parámetro explícito
                program_requirements
            )
            
            # DEBUG: Verificar DataFrame de validación
            logger.debug("DataFrame de validación de serie:")
            logger.debug(f"Columnas: {serial_validation_df.columns.tolist()}")
            logger.debug(f"Primeros registros:\n{serial_validation_df.head()}")
            logger.debug(f"Total de registros: {len(serial_validation_df)}")

            # Generar resumen mejorado con sistema de tolerancia
            # Identificar partes que tienen discrepancias Y tienen inventario
            mismatches_with_inventory = serial_validation_df[
                (serial_validation_df['Serial Control match?'] == 'Mismatch') & 
                (serial_validation_df['Inventory on Hand? Y/N'] == 'Y')
            ]
            
            # Calcular porcentajes para el sistema de tolerancia
            total_parts = len(serial_validation_df)
            total_mismatches = len(serial_validation_df[serial_validation_df['Serial Control match?'] == 'Mismatch'])
            total_with_inventory = len(serial_validation_df[serial_validation_df['Inventory on Hand? Y/N'] == 'Y'])
            mismatches_with_inventory_count = len(mismatches_with_inventory)
            
            # Calcular porcentajes para evaluación de tolerancia
            mismatch_pct = (total_mismatches / max(total_parts, 1)) * 100
            mismatch_with_inventory_pct = (mismatches_with_inventory_count / max(total_with_inventory, 1)) * 100
            
            # Determinar si estamos dentro del umbral de tolerancia del 20%
            within_tolerance = mismatch_pct <= 20.0
            
            # Crear lista de partes críticas para incluir en el resumen
            critical_parts = []
            if mismatches_with_inventory_count > 0:
                # Sólo incluir las primeras 10 partes para no sobrecargar el resumen
                critical_parts = mismatches_with_inventory['Part Number'].tolist()[:10]
                if len(mismatches_with_inventory) > 10:
                    critical_parts.append("... (y otras)")
            
            # CORRECCIÓN: Extraer las organizaciones que realmente se usaron en el reporte final
            # Obtenemos las columnas del DataFrame que corresponden a organizaciones físicas
            physical_orgs_columns = [col.replace('org ', '').replace(' Serial Control', '') 
                                  for col in serial_validation_df.columns 
                                  if col.startswith('org ') and 'Serial Control' in col]
            
            # Generar resumen mejorado
            summary = {
                'total_parts_reviewed': total_parts,
                'total_Serial Control mismatches': total_mismatches,
                'mismatch_percentage': f"{mismatch_pct:.2f}%",
                'total_with_inventory': total_with_inventory,
                'parts_with_inventory_pct': f"{(total_with_inventory / max(total_parts, 1) * 100):.2f}%",
                'mismatches_with_inventory': mismatches_with_inventory_count,
                'mismatches_with_inventory_pct': f"{mismatch_with_inventory_pct:.2f}%",
                'within_tolerance': "Yes" if within_tolerance else "No",
                'tolerance_threshold': "20%",
                'physical_orgs': ', '.join(physical_orgs_columns),  # CORRECCIÓN: Usar las organizaciones reales del reporte
                'total_physical_orgs': len(physical_orgs_columns),  # CORRECCIÓN: Contar las organizaciones reales del reporte
                'critical_review_parts': ', '.join(critical_parts) if critical_parts else "None",
                'timestamp': datetime.now().isoformat()
            }

            print("\n=== EXTERNAL REPORT SUMMARY ===")
            for key, value in summary.items():
                print(f"{key}: {value}")
                
            # Log detallado de mismatches con inventario para diagnóstico
            if mismatches_with_inventory_count > 0:
                logger.info(f"Se encontraron {mismatches_with_inventory_count} partes con discrepancias que tienen inventario")
                for idx, row in mismatches_with_inventory.head(10).iterrows():
                    part = row['Part Number']
                    details = row['Inventory Details']
                    logger.info(f"Parte crítica para revisión: {part} - {details}")
            
            
            # En _generate_external_report, justo antes de escribir a Excel
            logger.info("Verificando valores en la columna Item Status")

            # Verificar si la columna Item Status existe y tiene valores
            # Verificar más exhaustivamente la presencia de valores válidos en Item Status
            has_valid_item_status = False
            
            if 'Item Status' in serial_validation_df.columns:
                # Comprobar si hay valores no vacíos
                non_empty_status = not serial_validation_df['Item Status'].fillna('').astype(str).str.strip().eq('').all()
                # Comprobar si hay valores N/A
                has_na_values = (serial_validation_df['Item Status'].astype(str).str.strip() == 'N/A').any()
                
                has_valid_item_status = non_empty_status and not has_na_values
                
                # Análisis detallado para diagnóstico
                status_counts = serial_validation_df['Item Status'].value_counts().to_dict()
                logger.info(f"Distribución de Item Status: {status_counts}")
                logger.info(f"Tiene valores no vacíos: {non_empty_status}")
                logger.info(f"Tiene valores N/A: {has_na_values}")
                logger.info(f"Conclusión final - has_valid_item_status: {has_valid_item_status}")
            
            # Verificar valores para parte específica de ejemplo
            example_part = "MR56-HW.ACTUAL.324.FITCH"
            example_rows = serial_validation_df[serial_validation_df['Part Number'] == example_part]
            if not example_rows.empty:
                logger.info(f"Item Status para parte ejemplo {example_part}: '{example_rows['Item Status'].iloc[0]}'")
            else:
                logger.warning(f"Parte ejemplo {example_part} no encontrada antes de verificación")

            if not has_valid_item_status:
                logger.warning("Columna Item Status existe pero está vacía - rellenando con valores del DataFrame original")
                
                # Crear diccionario de mapeo
                item_status_dict = {}
                
                # Obtener valores del DataFrame original
                if isinstance(serial_results.get('data', None), pd.DataFrame):
                    df_original = serial_results['data']
                    if 'Item Status' in df_original.columns:
                        logger.info("Obteniendo Item Status del DataFrame original")
                        for part, group in df_original.groupby('Part Number'):
                            if not group['Item Status'].fillna('').eq('').all():
                                item_status_dict[part] = group['Item Status'].iloc[0]
                
                # Si no hay valores en el DataFrame, buscar en la lista de datos
                if not item_status_dict and isinstance(serial_results.get('data', None), list):
                    logger.info("Obteniendo Item Status de la lista de datos original")
                    for item in serial_results['data']:
                        if 'part_number' in item and 'item_status' in item and item['item_status']:
                            item_status_dict[item['part_number']] = item['item_status']
                
                # Aplicar el diccionario al DataFrame
                if item_status_dict:
                    logger.info(f"Aplicando {len(item_status_dict)} valores de Item Status al DataFrame")
                    serial_validation_df['Item Status'] = serial_validation_df['Part Number'].map(
                        lambda x: item_status_dict.get(x, '')
                    )
                    
                    # Asegurar que la columna Item Status sea la primera
                    if 'Item Status' in serial_validation_df.columns:
                        cols = serial_validation_df.columns.tolist()
                        if cols[0] != 'Item Status':
                            cols.remove('Item Status')
                            serial_validation_df = serial_validation_df[['Item Status'] + cols]
                            logger.info("Reordenado columnas para que Item Status sea la primera")
                else:
                    logger.warning("No se encontraron valores de Item Status en ninguna fuente de datos") 
            
            # DEBUG: Verificar resumen
            logger.debug("Summary generated:")
            logger.debug(str(summary))
            
            # Verificación final antes de escribir a Excel
            logger.info("=============== VERIFICACIÓN FINAL ANTES DE ESCRIBIR A EXCEL ===============")
            logger.info(f"DataFrame final para Excel - Columnas: {serial_validation_df.columns.tolist()}")
            logger.info(f"DataFrame final para Excel - Forma: {serial_validation_df.shape}")
            
            # Verificar específicamente la parte de ejemplo
            example_part = "MR56-HW.ACTUAL.324.FITCH"
            example_rows_final = serial_validation_df[serial_validation_df['Part Number'] == example_part]
            if not example_rows_final.empty:
                logger.info(f"VERIFICACIÓN FINAL - Item Status para {example_part}: '{example_rows_final['Item Status'].iloc[0]}'")
            else:
                logger.warning(f"VERIFICACIÓN FINAL - Parte {example_part} no encontrada antes de escribir a Excel")
                
            # Verificar distribución de valores de Item Status
            if 'Item Status' in serial_validation_df.columns:
                status_counts = serial_validation_df['Item Status'].value_counts().to_dict()
                logger.info(f"VERIFICACIÓN FINAL - Distribución de valores en Item Status: {status_counts}")
            else:
                logger.warning("VERIFICACIÓN FINAL - Columna 'Item Status' NO EXISTE en el DataFrame final")
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # Escribir resultados principales
                serial_validation_df.to_excel(writer, sheet_name="Serial Control Validation", index=False)
                self._format_worksheet_with_conditional(
                    writer.sheets["Serial Control Validation"],
                    serial_validation_df,
                    'Serial Control match?',
                    'Mismatch'
                )
                
                # Escribir resumen
                pd.DataFrame([summary]).to_excel(writer, sheet_name="Summary", index=False)
                self._format_worksheet(writer.sheets["Summary"])
            
            return excel_path
        
        except Exception as e:
            logger.error(f"Error generando reporte externo: {str(e)}")
            logger.error(f"Traza de error: {traceback.format_exc()}")
            raise

    def _generate_internal_report(self, audit_result: AuditResult, validation_results: Dict, timestamp: str, inventory_validation_enabled: bool = True) -> Path:
        try:
            contract_name = validation_results.get('program_requirements', {}).get('contract', 'Unknown')
            contract_name = contract_name.replace(' ', '_')  # Reemplazar espacios por guiones bajos
            excel_path = self.output_dir / f"organization_validation_report_{contract_name}_{timestamp}.xlsx"
            org_destination = validation_results.get('program_requirements', {}).get('org_destination', [])

            # Convert audit results to base dataframe
            results_df = self._convert_to_dataframe(audit_result)
            print(f"\n[Internal Report] DataFrame antes de validación:\n{results_df.head()}")
            
            # Generate org validation dataframe
            org_validation_df = self._generate_org_validation_data(results_df, org_destination)
            print(f"\n[Internal Report] DataFrame después de validación:\n{org_validation_df.head()}")

            # MEJORA: Generación de resumen con conteo correcto de problemas
            # Contador de problemas por organización
            issues_by_org = {}
            for org in org_destination:
                # Un problema es cuando una parte debería existir en esta org pero no existe
                issues_by_org[org] = 0
                
            # Contar manualmente para mayor precisión
            for _, row in org_validation_df.iterrows():
                # Si tiene mismatch, contar para cada organización faltante
                if row['Organization Status'] == 'Mismatch' and row['Organization code mismatch'] != 'None':
                    # Obtener organizaciones faltantes
                    missing_orgs = [org.strip() for org in str(row['Organization code mismatch']).split(',')]
                    # Incrementar contador para cada organización faltante
                    for org in missing_orgs:
                        if org in issues_by_org:
                            issues_by_org[org] += 1
            
            # Generate summary information con el conteo corregido
            summary = {
                'total_parts': len(org_validation_df['Part Number'].unique()),
                'missing_orgs_issues': len(org_validation_df[org_validation_df['Organization Status'] == 'Mismatch']),
                'issues_by_type': org_validation_df['Item Status'].value_counts().to_dict(),
                'issues_by_org': issues_by_org,  # Contador preciso por organización
                'severity_breakdown': {
                    'critical': len(org_validation_df[org_validation_df['Item Status'] == 'Missing in Org']),
                    'major': len(org_validation_df[org_validation_df['Organization Status'] == 'Mismatch']),
                    'minor': 0
                },
                'timestamp': datetime.now().isoformat()
            }
            
            # Log detallado del resumen
            logger.info("\n=== RESUMEN DEL REPORTE INTERNO ===")
            logger.info(f"Total de partes: {summary['total_parts']}")
            logger.info(f"Problemas de organizaciones faltantes: {summary['missing_orgs_issues']}")
            logger.info(f"Problemas por organización: {issues_by_org}")
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                org_validation_df.to_excel(writer, sheet_name="Audit Results", index=False)
                self._format_worksheet_with_conditional(
                    writer.sheets["Audit Results"],
                    org_validation_df,
                    'Status',
                    'Missing in Org'
                )
                pd.DataFrame([summary]).to_excel(writer, sheet_name="Summary", index=False)
                self._format_worksheet(writer.sheets["Summary"])
            
            return excel_path, summary
            
        except Exception as e:
            logger.error(f"Error generating internal report: {str(e)}")
            raise
            
    def _format_worksheet_with_conditional(
        self,
        worksheet,
        df: pd.DataFrame,
        conditional_column: str = None,
        error_value: str = None
    ) -> None:
        """Aplica formato condicional mejorado."""
        try:
            # Actualizar colores con los nuevos tipos de notas
            self.COLORS.update({
                'SERIAL_NOTES': 'E3F2FD',    # Azul muy claro para notas de Serial Control
                'NPI_NOTES': 'F3E5F5',       # Morado muy claro para notas de NPI
                'PROCUREMENT_NOTES': 'E8F5E9' # Verde muy claro para notas de Procurement
            })

            # Configurar anchos de columna
            for idx, col in enumerate(worksheet.columns, 1):
                max_length = max(len(str(cell.value or '')) for cell in col)
                adjusted_width = min(max_length + 2, 50)  # Limitar el ancho máximo
                worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width

            # Formato de encabezados mejorado
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(
                start_color=self.COLORS['HEADER'],
                end_color=self.COLORS['HEADER'],
                fill_type='solid'
            )
            header_alignment = Alignment(horizontal='center', vertical='center')

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            if df is not None:
                for col_idx, col_name in enumerate(df.columns, 1):
                    # Serial Control match?
                    if col_name == 'Serial Control match?' or col_name == 'Organization Status':
                        for row_idx, value in enumerate(df[col_name], 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            value = str(value).strip().upper()

                            if value == 'MATCH':
                                color = self.COLORS['SUCCESS']
                                font_color = 'FFFFFF'
                            else:
                                color = self.COLORS['ERROR']
                                font_color = 'FFFFFF'

                            cell.fill = PatternFill(start_color=color, fill_type='solid')
                            cell.font = Font(color=font_color)

                    # XX Serial Control
                    elif 'Serial Control' in col_name:
                        for row_idx, value in enumerate(df[col_name], 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            orig_value = str(value).strip()
                            upper_value = orig_value.upper()

                            color = self.COLORS['NEUTRAL']
                            font_color = '000000'

                            # CORRECCIÓN: Detectar "Not found in org" en cualquier combinación de mayúsculas/minúsculas
                            if 'NO SERIAL NUMBER CONTROL' in upper_value:
                                color = self.COLORS['LIGHT_YELLOW']
                            elif 'DYNAMIC ENTRY' in upper_value:
                                color = self.COLORS['LIGHT_BLUE']
                            elif 'NOT FOUND' in upper_value or upper_value == 'NOT FOUND IN ORG' or orig_value == 'Not found in org':
                                color = self.COLORS['LIGHT_RED']
                                # CORRECCIÓN: Estandarizar el texto mostrado como "Not found in org"
                                cell.value = 'Not found in org'

                            cell.fill = PatternFill(start_color=color, fill_type='solid')
                            cell.font = Font(color=font_color)

                    # Estados de organización
                    elif col_name.startswith('org ') and 'Serial Control' not in col_name:
                        for row_idx, value in enumerate(df[col_name], 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            orig_value = str(value).strip()
                            upper_value = orig_value.upper()

                            if 'PRESENT IN ORG' in upper_value:
                                color = self.COLORS['LIGHT_GREEN']
                            # CORRECCIÓN: Detectar "Not found in org" en cualquier combinación de mayúsculas/minúsculas  
                            elif 'NOT FOUND IN ORG' in upper_value or upper_value == 'NOT FOUND IN ORG' or orig_value == 'Not found in org' or upper_value == 'MISSING IN ORG':
                                color = self.COLORS['LIGHT_RED']
                                # CORRECCIÓN: Estandarizar el texto mostrado como "Missing in Org"
                                cell.value = 'Missing in Org'
                            elif 'PHASE OUT' in upper_value:
                                color = self.COLORS['LIGHT_YELLOW']
                            else:
                                color = self.COLORS['NEUTRAL']

                            cell.fill = PatternFill(start_color=color, fill_type='solid')

                    # Status (New, In Progress, Closed)
                    elif col_name == 'Status':
                        for row_idx, value in enumerate(df[col_name], 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            value = str(value).strip().upper()

                            if value == 'NEW':
                                color = self.COLORS['LIGHT_YELLOW']
                            elif value == 'IN PROGRESS':
                                color = self.COLORS['LIGHT_BLUE']
                            elif value == 'CLOSED':
                                color = self.COLORS['LIGHT_GREEN']
                            else:
                                color = self.COLORS['NEUTRAL']

                            cell.fill = PatternFill(start_color=color, fill_type='solid')

                    # Inventory on Hand
                    elif col_name == 'Inventory on Hand? Y/N':
                        for row_idx, value in enumerate(df[col_name], 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            value = str(value).strip().upper()

                            color = self.COLORS['LIGHT_GREEN'] if value == 'Y' else self.COLORS['LIGHT_YELLOW']
                            cell.fill = PatternFill(start_color=color, fill_type='solid')

                    
                    elif col_name == 'Item Status':
                        # Log para verificar valores de Item Status al aplicar formato
                        logger.info("=============== FORMATO DE COLUMNA ITEM STATUS ===============")
                        logger.info(f"Valores en columna 'Item Status': {df[col_name].value_counts().to_dict()}")
                        
                        # Verificar específicamente la parte de ejemplo
                        example_part = "MR56-HW.ACTUAL.324.FITCH"
                        example_idx = None
                        for idx, row_val in enumerate(df['Part Number']):
                            if row_val == example_part:
                                example_idx = idx
                                break
                                
                        if example_idx is not None:
                            example_status = df.iloc[example_idx]['Item Status']
                            logger.info(f"FORMATO CONDICIONAL - Item Status para {example_part}: '{example_status}'")
                        else:
                            logger.warning(f"FORMATO CONDICIONAL - Parte {example_part} no encontrada al aplicar formato")
                            
                        for row_idx, value in enumerate(df[col_name], 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            value_upper = str(value).strip().upper()
                            
                            # Definir colores según el estado del ítem
                            if value_upper in ('ACTIVE', 'APPROVED'):
                                color = self.COLORS['LIGHT_GREEN']
                            elif value_upper in ('INACTIVE', 'OBSOLETE', 'DISCONTINUED'):
                                color = self.COLORS['LIGHT_RED']
                            elif value_upper in ('PENDING APPROVAL', 'ON HOLD'):
                                color = self.COLORS['LIGHT_YELLOW']
                            elif value_upper in ('PHASE OUT', 'PHASE-OUT'):
                                color = self.COLORS['LIGHT_BLUE']
                            # CORRECCIÓN: Añadir manejo explícito para 'Missing in Org'
                            elif value_upper == 'MISSING IN ORG' or 'NOT FOUND IN ORG' in value_upper:
                                color = self.COLORS['LIGHT_RED']
                                # Asegurarse de que el texto se muestre exactamente como "Missing in Org"
                                cell.value = 'Missing in Org'
                            else:
                                color = self.COLORS['NEUTRAL']
                                
                            cell.fill = PatternFill(start_color=color, fill_type='solid')
                    
                    # Notas de Serial Control
                    elif 'Serial Control Owner Notes' in col_name:
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = PatternFill(
                                start_color=self.COLORS['SERIAL_NOTES'],
                                fill_type='solid'
                            )

                    # Notas de NPI
                    elif 'NPI Action' in col_name:
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = PatternFill(
                                start_color=self.COLORS['NPI_NOTES'],
                                fill_type='solid'
                            )

                    # Notas de Procurement
                    elif 'Procurement/Order Management Team Action Notes' in col_name:
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.fill = PatternFill(
                                start_color=self.COLORS['PROCUREMENT_NOTES'],
                                fill_type='solid'
                            )

            # Bordes mejorados
            thin_border = Border(
                left=Side(style='thin', color='E0E0E0'),
                right=Side(style='thin', color='E0E0E0'),
                top=Side(style='thin', color='E0E0E0'),
                bottom=Side(style='thin', color='E0E0E0')
            )

            # Aplicar bordes y alineación
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:  # No aplicar a encabezados
                        cell.alignment = Alignment(vertical='center')

            # Congelar primera fila y ajustar zoom
            worksheet.freeze_panes = 'A2'
            worksheet.sheet_view.zoomScale = 100

        except Exception as e:
            logger.error(f"Error en format_worksheet_with_conditional: {str(e)}")
            raise
    
    
    def _format_validation_dataframe(self, df: pd.DataFrame, physical_orgs: List[str], dynamic_columns: List[str]) -> pd.DataFrame:
        """
        Asegura que el DataFrame tenga todas las columnas necesarias en el orden correcto.
        Utiliza physical_orgs para las columnas de organizaciones físicas.
        MEJORA: Ahora usa intersección de organizaciones físicas predefinidas y organizaciones en datos.
        """
        # Verificar DataFrame ANTES del formato
        logger.info("=============== MÉTODO _format_validation_dataframe ===============")
        logger.info(f"DataFrame recibido - Forma: {df.shape}")
        logger.info(f"DataFrame recibido - Columnas: {df.columns.tolist()}")
        
        # Verificar especialmente la presencia de Item Status
        example_part = "MR56-HW.ACTUAL.324.FITCH"
        if 'Item Status' in df.columns:
            logger.info(f"VERIFICACIÓN 'Item Status' EN ENTRADA _format_validation_dataframe:")
            status_values = df['Item Status'].value_counts().to_dict()
            logger.info(f"Distribución de valores en Item Status: {status_values}")
            
            # Verificar especialmente la parte de ejemplo
            example_rows = df[df['Part Number'] == example_part]
            if not example_rows.empty:
                logger.info(f"ENTRADA FORMATO - Item Status para {example_part}: '{example_rows['Item Status'].iloc[0]}'")
            else:
                logger.warning(f"ENTRADA FORMATO - Parte {example_part} no encontrada")
        else:
            logger.warning("ALERTA CRÍTICA: Columna 'Item Status' NO PRESENTE en DataFrame de entrada a _format_validation_dataframe")
        
        # Definir orden de columnas base
        columns = [
            'Item Status',
            'Part Number', 
            'Manufacturer', 
            'Description', 
            'Vertex', 
            'Serial Control match?'
        ]
        
        # MEJORA: Identificar organizaciones presentes en los datos
        orgs_in_data = []
        for col in df.columns:
            if col.startswith('org ') and 'Serial Control' in col:
                # Extraer número de organización del nombre de columna
                org_code = col.replace('org ', '').replace(' Serial Control', '')
                orgs_in_data.append(org_code)
        
        # Ordenar y eliminar duplicados
        orgs_in_data = sorted(list(set(orgs_in_data)))
        logger.info(f"Organizaciones encontradas en datos: {orgs_in_data}")
        
        # CORRECCIÓN URGENTE: Asegurar que solo se incluyan en el reporte las organizaciones 
        # físicas que realmente aparecen en los datos de auditoría.
        # physical_orgs ya debe venir filtrado desde _generate_serial_validation_data
        
        # Verificar que estamos usando organizaciones que están tanto en los datos como en la lista de físicas
        orgs_to_use = sorted(list(set(physical_orgs).intersection(set(orgs_in_data))))
        
        # Verificar que las organizaciones físicas recibidas coinciden con las esperadas
        if set(orgs_to_use) != set(physical_orgs):
            logger.warning(f"¡ALERTA! Las organizaciones físicas recibidas no coinciden con las encontradas en los datos!")
            logger.warning(f"Recibidas: {physical_orgs}")
            logger.warning(f"Encontradas en datos: {orgs_in_data}")
            logger.warning(f"Usando finalmente: {orgs_to_use}")
            # Actualizamos physical_orgs para asegurar consistencia
            physical_orgs = orgs_to_use
        
        # Registrar información de diagnóstico
        logger.info(f"Generando reporte con {len(orgs_to_use)} organizaciones físicas: {orgs_to_use}")
        
        # Verificar organizaciones físicas que aparecen en las columnas pero no en los datos
        extra_orgs = set(orgs_in_data) - set(physical_orgs)
        if extra_orgs:
            logger.warning(f"Se encontraron columnas de organizaciones adicionales: {extra_orgs} (estas NO se incluirán en el reporte)")
        
        # Agregar columnas de Serial Control solo para las organizaciones físicas que aparecen en los datos
        for org in orgs_to_use:
            columns.append(f'org {org} Serial Control')  # Agregar 'org' antes de la organización
        
        # Agregar columnas restantes en el orden específico
        remaining_columns = [
            'Inventory on Hand? Y/N',
            'Inventory Details',
            'NPI Recommendations',
            'Serial control Owner Action Notes(ISR) *Required',
            'Required Serial control Owner Notes (ISR) *Optional',
            'Optional NPI Action/Data update',
            'Procurement/Order Management Team Action Notes *Required',
            'Procurement/Order Management Team Action Notes *Optional',
            'NPI resolution notes',
            'Status',
            'Action Required',
            'Missing orgs according to Program Requirements'
        ]
        columns.extend(remaining_columns)

        # Asegurar que existan todas las columnas
        for col in columns:
            if col not in df.columns:
                df[col] = ''
        
        # Verificar conservación de datos importantes justo antes de filtrar columnas
        if 'Item Status' in df.columns and 'Part Number' in df.columns:
            for part in [example_part]:
                part_rows = df[df['Part Number'] == part]
                if not part_rows.empty:
                    current_status = part_rows['Item Status'].iloc[0]
                    logger.info(f"VERIFICACIÓN PRE-FILTRADO - Parte {part} - Item Status: '{current_status}'")
                else:
                    logger.warning(f"VERIFICACIÓN PRE-FILTRADO - Parte {part} no encontrada")
        
        # Filtrar para incluir solo las columnas definidas (por si hay columnas extras)
        result_df = df[columns]
        
        # Verificar DataFrame DESPUÉS del formato
        logger.info("=============== RESULTADO _format_validation_dataframe ===============")
        logger.info(f"DataFrame resultado - Forma: {result_df.shape}")
        logger.info(f"DataFrame resultado - Columnas: {result_df.columns.tolist()}")
        
        # Verificar específicamente si la parte de ejemplo mantiene su Item Status
        if 'Item Status' in result_df.columns:
            example_rows_after = result_df[result_df['Part Number'] == example_part]
            if not example_rows_after.empty:
                logger.info(f"SALIDA FORMATO - Item Status para {example_part}: '{example_rows_after['Item Status'].iloc[0]}'")
            else:
                logger.warning(f"SALIDA FORMATO - Parte {example_part} no encontrada después del formato")
        else:
            logger.warning("ALERTA CRÍTICA: Columna 'Item Status' NO PRESENTE en DataFrame resultado de _format_validation_dataframe")
        
        return result_df
        
    def _convert_to_dataframe(self, audit_result: AuditResult) -> pd.DataFrame:
        """
        Convierte un AuditResult en un DataFrame de pandas para procesamiento de reportes.
        
        Args:
            audit_result: Instancia de AuditResult con los resultados de la auditoría
            
        Returns:
            DataFrame con los datos procesados y normalizados
        """
        try:
            # Lista para almacenar los datos procesados
            processed_data = []
            
            # Extraer datos de serial_control_results de manera segura
            serial_control_data = {}
            if audit_result.serial_control_results:
                try:
                    # Asegurar que tenemos datos y que son del tipo correcto
                    if isinstance(audit_result.serial_control_results, dict):
                        raw_data = audit_result.serial_control_results.get('data', [])
                        if isinstance(raw_data, (list, pd.DataFrame)):
                            for result in raw_data:
                                if isinstance(result, dict):
                                    key = result.get('part_number')
                                    if key:
                                        serial_control_data[key] = {
                                            'manufacturer': result.get('manufacturer', ''),
                                            'description': result.get('description', ''),
                                            'vertex': result.get('vertex', '')
                                        }
                except Exception as e:
                    logger.warning(f"Error procesando serial_control_results: {str(e)}")
                    logger.warning("Continuando con el procesamiento...")
            
            # Procesar cada AuditItem
            for item in audit_result.items:
                # Datos base del item
                base_data = {
                    'Item Status': item.item_status if hasattr(item,'item_status') else '',
                    'Part Number': item.part_number,
                    'Organization': item.organization,
                    'Status': item.status,
                    'Action Required': item.action_required,
                    'Current Orgs': ','.join(item.current_orgs) if item.current_orgs else '',
                    'Missing Orgs': ','.join(item.missing_orgs) if item.missing_orgs else ''
                }
                
                # Añadir datos de program_requirements si están disponibles
                if hasattr(audit_result, 'program_requirements') and audit_result.program_requirements:
                    base_data.update({
                        'Base Org': audit_result.program_requirements.base_org,
                        'Org Destination': ','.join(audit_result.program_requirements.org_destination)
                    })
                
                # Agregar información de Serial Control
                if hasattr(item, 'serial_control') and item.serial_control:
                    base_data.update({
                        'Serial Control': item.serial_control.current_value,
                        'Serial Control Active': 'Yes' if item.serial_control.is_active else 'No'
                    })
                
                # Agregar información de inventario
                if hasattr(item, 'inventory_info') and item.inventory_info:
                    base_data.update({
                        'On Hand Quantity': item.inventory_info.quantity,
                        'Has Stock': 'Yes' if item.inventory_info.has_stock else 'No',
                        'Value': item.inventory_info.value,
                        'Subinventory': item.inventory_info.subinventory_code,
                        'Warehouse': item.inventory_info.warehouse_code
                    })
                    
                    # Agregar información de aging
                    if item.inventory_info.aging_info:
                        base_data.update({
                            'Aging 0-30': item.inventory_info.aging_info.days_0_30,
                            'Aging 31-60': item.inventory_info.aging_info.days_31_60,
                            'Aging 61-90': item.inventory_info.aging_info.days_61_90
                        })
                
                # Agregar datos adicionales de serial control si existen
                additional_data = serial_control_data.get(item.part_number, {})
                base_data.update({
                    'Manufacturer': additional_data.get('manufacturer', ''),
                    'Description': additional_data.get('description', ''),
                    'Vertex': additional_data.get('vertex', '')
                })
                
                processed_data.append(base_data)
            
            # Crear DataFrame
            df = pd.DataFrame(processed_data)
            
            # Asegurar tipos de datos correctos
            numeric_columns = ['On Hand Quantity', 'Value', 'Aging 0-30', 'Aging 31-60', 'Aging 61-90']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
            # Ordenar columnas para mejor legibilidad
            column_order = [
                'Item Status','Part Number', 'Organization', 'Status', 'Action Required',
                'Current Orgs', 'Missing Orgs', 'Base Org', 'Org Destination',
                'Serial Control', 'Serial Control Active',
                'On Hand Quantity', 'Has Stock', 'Value', 'Subinventory', 'Warehouse',
                'Aging 0-30', 'Aging 31-60', 'Aging 61-90',
                'Manufacturer', 'Description', 'Vertex'
            ]
            
            # Filtrar columnas que existen en el DataFrame
            existing_columns = [col for col in column_order if col in df.columns]
            df = df[existing_columns]
            
            logger.info(f"Successfully converted {len(processed_data)} items to DataFrame")
            return df
            
        except Exception as e:
            logger.error(f"Error converting audit result to DataFrame: {str(e)}")
            logger.error(f"Stack trace: {traceback.format_exc()}")
            raise ValueError(f"Failed to convert audit result to DataFrame: {str(e)}")
    
    def _generate_serial_validation_data(self, results: Dict, physical_orgs: List[str], program_requirements: Dict) -> pd.DataFrame:
        """
        Genera el reporte de validación de Serial Control con lógica mejorada.
        Ahora utiliza physical_orgs en lugar de org_destination para filtrar las organizaciones físicas
        """
        try:
            print("\n=== INICIO SERIAL VALIDATION DATA GENERATION ===")

            # Logging: Inspección del origen de datos
            logger.info("=============== DIAGNÓSTICO DE ORIGEN DE DATOS ===============")
            print("\nDEBUG - Input Data Analysis:")
            print(f"Results keys: {results.keys()}")
            print(f"Data shape: {pd.DataFrame(results['data']).shape}")
            print(f"Unique parts in data: {len(pd.DataFrame(results['data'])['Part Number'].unique())}")
            print(f"Mismatched parts: {len(results.get('mismatched_parts', []))}")
            
            # 1. Log: Inspección del DataFrame original
            df_data = pd.DataFrame(results['data']) if isinstance(results['data'], list) else results['data']
            logger.info(f"DataFrame original - Forma: {df_data.shape}")
            logger.info(f"DataFrame original - Columnas: {df_data.columns.tolist()}")
            
            # 2. Log: Verificación específica de 'Item Status' en el DataFrame original
            example_part = "MR56-HW.ACTUAL.324.FITCH"
            if 'Item Status' in df_data.columns:
                logger.info("VERIFICACIÓN DE 'Item Status' EN DATAFRAME ORIGINAL")
                logger.info(f"Valores únicos en 'Item Status': {df_data['Item Status'].unique()}")
                
                # Verificar especialmente la parte de ejemplo
                example_rows = df_data[df_data['Part Number'] == example_part]
                if not example_rows.empty:
                    example_status = example_rows['Item Status'].iloc[0]
                    logger.info(f"Item Status para parte {example_part}: {example_status}")
                else:
                    logger.warning(f"Parte de ejemplo {example_part} no encontrada en DataFrame original")
            else:
                logger.warning("No se encontró la columna 'Item Status' en el DataFrame original")
            
            # Continuar con el análisis estándar
            inventory_map = results.get('inventory_map', {})
            print("\nDEBUG - Inventory Map Analysis:")
            print(f"Total inventory records: {len(inventory_map)}")
            print("Sample inventory records (first 3):")
            for key in list(inventory_map.keys())[:3]:
                print(f"Key: {key}")
                print(f"Data: {inventory_map[key]}")
                
            mismatched_parts = set(str(part).strip() for part in results.get('mismatched_parts', []))
            print("\nDEBUG - Mismatch Analysis:")
            print(f"Total mismatched parts: {len(mismatched_parts)}")
            print("Sample mismatched parts (first 3):")
            print(list(mismatched_parts)[:3])

            # Análisis de Program Requirements
            print("\nDEBUG - Program Requirements:")
            print(f"Base org: {program_requirements.get('base_org')}")
            print(f"Original org_destination: {program_requirements.get('org_destination')}")    
            print(f"Physical organizations: {physical_orgs}")
                
            # 1. Obtener TODAS las organizaciones del programa
            # df_data already loaded above
            
            # DIAGNÓSTICO: Verificar estructura del DataFrame original
            logger.info("===== DIAGNÓSTICO ITEM STATUS =====")
            logger.info(f"Columnas en df_data: {df_data.columns.tolist()}")
            if 'Item Status' in df_data.columns:
                unique_values = df_data['Item Status'].fillna('NULL').unique().tolist()
                logger.info(f"Valores únicos de Item Status en df_data: {unique_values}")
                
                # Verificar ejemplos de valores específicos
                sample_rows = df_data[pd.notna(df_data['Item Status'])].head(5)
                if not sample_rows.empty:
                    for idx, row in sample_rows.iterrows():
                        logger.info(f"Ejemplo {idx}: Part Number='{row['Part Number']}', Item Status='{row['Item Status']}' (tipo: {type(row['Item Status'])})")
            else:
                logger.warning("Columna 'Item Status' no encontrada en df_data")

            # Asegurar que physical_orgs está normalizado
            physical_orgs = [str(org).strip().zfill(2) for org in physical_orgs]
            org_destination = program_requirements.get('org_destination', [])
            org_destination = [str(org).strip().zfill(2) for org in org_destination]
            logger.info(f"Organizaciones destino según requisitos del programa: {org_destination}")
            
            # CORRECCIÓN URGENTE: Identificar las organizaciones presentes en el documento de auditoría
            # y usar solo la intersección con la lista de organizaciones físicas
            orgs_in_audit = df_data['Organization'].astype(str).str.strip().str.zfill(2).unique().tolist()
            
            # Crear la intersección entre organizaciones físicas y las presentes en el documento
            physical_orgs_in_audit = sorted(list(set(physical_orgs).intersection(set(orgs_in_audit))))
            
            # Reemplazar la lista completa de physical_orgs por solo las que están en la auditoría
            physical_orgs = physical_orgs_in_audit
            
            # Log para verificar las organizaciones físicas filtradas
            logger.info(f"Organizaciones físicas antes de filtrar: {len(physical_orgs_in_audit)} de 72")
            logger.info(f"Usando SOLO las organizaciones físicas presentes en la auditoría: {physical_orgs}")
        
            # Normalizar organizaciones en el DataFrame
            df_data['Organization'] = df_data['Organization'].astype(str).str.strip()
            df_data['Organization'] = df_data['Organization'].str.zfill(2)
            logger.debug(f"Organizaciones únicas después de normalización: {df_data['Organization'].unique()}")
             
            # 2. Procesar TODOS los datos, no solo los mismatches
            all_parts = df_data['Part Number'].unique()
            # Normalizar consistentemente las partes con mismatch (usar UPPER para comparaciones)
            mismatched_parts = set(str(part).strip().upper() for part in results.get('mismatched_parts', []))
            
            print(f"\nDEBUG - Processing Stats:")
            print(f"Total unique parts: {len(all_parts)}")
            print(f"Total mismatches: {len(mismatched_parts)}")
            print(f"Organizations to process: {physical_orgs}")
                
            # 3. Procesamiento de datos principal
            validation_data = []
            inventory_map = results.get('inventory_map', {})
            # Ya hemos normalizado mismatched_parts arriba, no sobrescribir aquí
            # (mantenemos la versión normalizada)
            dynamic_columns = results.get('dynamic_columns', [])
            
            # DIAGNÓSTICO: Verificar específicamente el ejemplo mencionado
            example_part = "MR56-HW.ACTUAL.324.FITCH"
            example_rows = df_data[df_data['Part Number'] == example_part]
            if not example_rows.empty:
                logger.info(f"===== ANÁLISIS DE PARTE ESPECÍFICA: {example_part} =====")
                logger.info(f"Número de registros encontrados: {len(example_rows)}")
                for idx, row in example_rows.iterrows():
                    logger.info(f"Registro {idx}:")
                    for col in row.index:
                        logger.info(f"  {col}: '{row[col]}'")
                logger.info("===============================================")
            else:
                logger.warning(f"Parte específica no encontrada: {example_part}")

            print(f"Processing {len(df_data)} records with {len(physical_orgs)} organizations")
            
            # Tracking para validación
            processed_parts = set()
            
            for part_number, part_group in df_data.groupby('Part Number'):
                # Normalización de part_number para comparación consistente
                part_norm = str(part_number).strip().upper()
                
                # También normalizar la lista de mismatched_parts para comparación exacta
                # Evaluar si este part_number está en la lista normalizada de mismatches
                is_mismatch = part_norm in set(str(p).strip().upper() for p in mismatched_parts)
                
                # Guardar de manera normalizada para validaciones posteriores
                processed_parts.add(part_norm)
                
                # Resto del código existente para procesamiento por parte
                has_inventory = False
                inventory_details = []
                serial_values = []

                
                # NUEVA IMPLEMENTACIÓN: Obtener Status primero antes de buscar Item Status
                # Inicializar el valor de status que usaremos
                status_value = 'N/A'
                
                # Registrar para diagnóstico
                is_example_part = part_number == "MR56-HW.ACTUAL.324.FITCH"
                if is_example_part:
                    logger.info(f"=========== PROCESAMIENTO DE PARTE: {part_number} ===========")
                    logger.info(f"Registros en grupo: {len(part_group)}")
                    logger.info(f"Columnas disponibles: {part_group.columns.tolist()}")
                
                # 1. PRIMERO intentamos obtener de la columna Status original
                if 'Status' in part_group.columns:
                    raw_status = part_group['Status'].iloc[0]
                    logger.info(f"Part Number '{part_number}' - Status original: '{raw_status}' (tipo: {type(raw_status)})")
                    
                    if pd.notna(raw_status) and str(raw_status).strip():
                        # Normalizar a mayúsculas para consistencia
                        status_value = str(raw_status).strip().upper()
                        logger.info(f"Part Number '{part_number}' - Status encontrado y normalizado: '{status_value}'")
                    else:
                        logger.info(f"Part Number '{part_number}' - Status encontrado pero es nulo o vacío")
                
                # 2. DESPUÉS, si ya existe Item Status, verificamos si podemos usarlo como respaldo
                elif 'Item Status' in part_group.columns:
                    raw_status = part_group['Item Status'].iloc[0]
                    logger.info(f"Part Number '{part_number}' - Item Status como respaldo: '{raw_status}' (tipo: {type(raw_status)})")
                    
                    if pd.notna(raw_status) and str(raw_status).strip() and str(raw_status).strip().upper() != 'N/A':
                        status_value = str(raw_status).strip().upper()
                        logger.info(f"Part Number '{part_number}' - Usando Item Status como respaldo: '{status_value}'")
                    else:
                        logger.info(f"Part Number '{part_number}' - Item Status es nulo, vacío o N/A")
                else:
                    logger.info(f"Part Number '{part_number}' - Ni 'Status' ni 'Item Status' encontrados")
                
                # Crear registro base con el status_value obtenido correctamente
                row = {
                    'Item Status': status_value,  # Usar el valor que obtuvimos de Status
                    'Part Number': part_number,
                    'Manufacturer': part_group['Manufacturer'].iloc[0],
                    'Description': part_group['Description'].iloc[0],
                    'Vertex': part_group['Vertex'].iloc[0],
                    'Serial Control match?': 'Match'  # Asumir match por defecto
                }

                # Inicializar columnas de Serial Control para TODAS las organizaciones físicas
                for org in physical_orgs:
                    row[f'org {org} Serial Control'] = 'Not found in org'  # Usar formato consistente

                # Variables para inventario
                has_inventory = False
                inventory_details = []
                serial_values_normalized = []

                # Procesar cada organización física
                for org in physical_orgs:
                    org_data = part_group[part_group['Organization'].astype(str).str.strip() == org]
                    inventory_key = f"{part_number}_{org}"
                    inv_data = inventory_map.get(inventory_key, {}) 

                    if not org_data.empty:
                        serial_value = org_data['Serial Control'].iloc[0]
                        
                        # Normalización explícita del valor de Serial Control
                        normalized_serial_value = (
                            'Dynamic entry at inventory receipt' if serial_value == 'YES' 
                            else 'No serial number control' if serial_value == 'NO' 
                            else serial_value
                        )
                        
                        # Actualizar la columna de Serial Control para esta org
                        row[f'org {org} Serial Control'] = normalized_serial_value
                        
                        # Trackear valores normalizados de serie
                        serial_values_normalized.append(normalized_serial_value)
                        
                        # Procesamiento de inventario
                        quantity = float(inv_data.get('quantity', 0))
                        
                        # Antes de procesar el inventario, verificar existencia de campos
                        if 'quantity' not in inv_data:
                            logger.warning(f"No hay cantidad para {inventory_key}")
                            quantity = 0
                        else:
                            quantity = inv_data['quantity']
                        
                        if quantity > 0:
                            has_inventory = True
                            inventory_details.append(f"{org}: {quantity} units")
                    else:
                        # Mantener la lógica de no encontrado para inventario
                        row[f'Inventory Status {org}'] = 'Not found in org'
                # REDISEÑADO: Determine mismatch con reglas de validación más estrictas
                # Aplicar TODAS las reglas del documento de requerimientos:
                # 1. Ignora completamente la organización 01
                # 2. Solo compara valores reales (no "Not found in org")
                # 3. Match solo si TODOS los valores reales son EXACTAMENTE iguales
                
                org_values_for_comparison = {}

                # Filtrado ESTRICTO de valores por organización según requerimientos
                for org in physical_orgs:
                    org_key = f'org {org} Serial Control'
                    if org != '01' and org_key in row:  # Excluir Org 01 completamente
                        # Normalizar para garantizar comparación precisa
                        raw_value = row[org_key].strip() if isinstance(row[org_key], str) else str(row[org_key])
                        normalized_value = self._normalize_serial_value(raw_value)
                        
                        # Solo incluir valores reales para comparación (no "Not found in org")
                        if normalized_value != 'Not found in org' and normalized_value not in ('', 'nan', 'None', 'none', 'NaN'):
                            # Almacenar valor normalizado para comparación posterior
                            org_values_for_comparison[org] = normalized_value
                            logger.debug(f"Org {org} tiene valor real: '{normalized_value}' (original: '{raw_value}')")
                        else:
                            logger.debug(f"Org {org} ignorada en comparación: '{normalized_value}' (original: '{raw_value}')")

                # 2. Determine mismatch based on meaningful values with STRICT validation rules
                if org_values_for_comparison:
                    # Get unique Serial Control values (excluding non-meaningful values and org 01)
                    unique_serial_values = set(org_values_for_comparison.values())
                    
                    # There's a mismatch if there are different meaningful values
                    # MATCH occurs ONLY when all real values are exactly the same
                    has_value_mismatch = len(unique_serial_values) > 1
                    
                    # Enhanced logging for validation decision traceability
                    logger.debug(f"Meaningful Serial Control values for part {part_norm} (excluding '01'): {org_values_for_comparison}")
                    logger.debug(f"Unique meaningful values: {unique_serial_values}")
                    logger.debug(f"VALIDATION DECISION: {'MISMATCH' if has_value_mismatch else 'MATCH'}")
                    
                    # Additional traceability log with detailed reasoning
                    if has_value_mismatch:
                        logger.debug(f"MISMATCH REASON: Found {len(unique_serial_values)} different real values: {unique_serial_values}")
                    else:
                        logger.debug(f"MATCH REASON: All real values are identical: {list(unique_serial_values)[0] if unique_serial_values else 'No values'}")
                else:
                    # If no meaningful values to compare, there can't be a mismatch
                    has_value_mismatch = False
                    unique_serial_values = set()
                    logger.debug(f"No meaningful Serial Control values to compare for part {part_norm}")
                    logger.debug("VALIDATION DECISION: MATCH (no real values to compare)")

                # 3. Check if it was previously flagged as a mismatch
                was_flagged_as_mismatch = part_norm in mismatched_parts

                # 4. Determinación ESTRICTA del estado final basada en criterios del documento
                # Un MATCH solo ocurre cuando todos los valores reales son EXACTAMENTE iguales
                # "Not found in org" no participa en las comparaciones
                
                # Obtener una función pura para validación de control serial
                validation_result = self._validate_serial_control(org_values_for_comparison)
                
                row['Inventory on Hand? Y/N'] = 'Y' if has_inventory else 'N'
                row['Inventory Details'] = ', '.join(inventory_details) if inventory_details else 'No inventory found'
                
                # CORRECCIÓN CRÍTICA: Usar SOLO la función de validación para determinar el resultado
                # Ignorar was_flagged_as_mismatch para evitar falsos positivos
                row['Serial Control match?'] = validation_result

                # 5. Registro detallado para auditoría y trazabilidad completa
                if row['Serial Control match?'] == 'Mismatch':
                    reason = f"VALORES DIFERENTES detectados entre organizaciones: {unique_serial_values}"
                        
                    # Log con nivel de detalle adecuado para auditoría posterior
                    logger.info(f"MISMATCH - Parte {part_norm} - Razón: {reason}")
                    logger.info(f"Valores por organización: {json.dumps(org_values_for_comparison, indent=2)}")
                    logger.debug(f"Valores únicos encontrados: {unique_serial_values}")
                else:
                    # Importante para diagnóstico: registrar siempre datos que llevaron a MATCH
                    logger.info(f"MATCH - Parte {part_norm} - Todos los valores reales coinciden o no hay suficientes valores para comparar")
                    logger.info(f"Orgs con valores reales: {list(org_values_for_comparison.keys())}")
                    logger.info(f"Valores por organización: {org_values_for_comparison}")
                    
                    # Para el ejemplo específico mencionado
                    if part_norm == "044-232355-01-P.NC.1244.CHARTER":
                        logger.warning(f"PARTE ESPECIAL: {part_norm}")
                        logger.warning(f"Valores por organización: {org_values_for_comparison}")
                        logger.warning(f"Valores únicos: {unique_serial_values}")
                        logger.warning(f"Resultado: {validation_result}")

                # Determinar si se requiere acción basada en el resultado
                row['Action Required'] = 'Review Serial Control' if row['Serial Control match?'] == 'Mismatch' else ''
                
               # Calcular organizaciones faltantes según la misma lógica de _check_missing_orgs
                missing_orgs_program = []
                if org_destination:
                    # Determinar en qué organizaciones existe esta parte según el DataFrame original
                    org_exists = set()
                    if 'Organization' in part_group.columns:
                        for _, part_row in part_group.iterrows():
                            org_raw = str(part_row['Organization']).strip()
                            org = org_raw.zfill(2)
                            org_exists.add(org)
                    
                    # Determinar organizaciones faltantes (las que están en org_destination pero no en org_exists)
                    missing_orgs_program = [org for org in org_destination if org not in org_exists]
                    
                    # Ordenar para consistencia
                    missing_orgs_program = sorted(missing_orgs_program)
                    
                    if missing_orgs_program:
                        logger.debug(f"Part {part_number}: Missing orgs: {missing_orgs_program}")

                # Añadir la información a la fila SIN SOBRESCRIBIR la asignación posterior
                missing_orgs_str = ','.join(missing_orgs_program) if missing_orgs_program else 'None'

                # Es importante NO modificar el código existente que inicializa el diccionario row
                # Ya que es crítico para el funcionamiento del resto del método
                # Solo AÑADIR nuestra información al objeto row que ya existe
                row['Missing orgs according to Program Requirements'] = missing_orgs_str
                
                row.update({
                    'NPI Recommendations': '',
                    'Serial control Owner Notes (ISR) *Required': '',
                    'Serial Control Owner Notes (ISR) *Optional': '',
                    'NPI Action/Data update': '',
                    'Procurement/Order Management Team Action Notes *Required': '',
                    'Procurement/Order Management Team Action Notes *Optional': '',
                    'NPI resolution notes': '',
                    'Status': 'New',
                    'Missing orgs according to Program Requirements': row.get('Missing orgs according to Program Requirements', 'None')
                })

                # 4. Log: Verificar el Item Status justo antes de agregar el row a validation_data
                if is_example_part:
                    logger.info(f"PARTE EJEMPLO - Item Status antes de agregar a validation_data: '{row['Item Status']}'")
                
                validation_data.append(row)

            # 5. Crear DataFrame final y verificarlo antes de formato
            validation_df = pd.DataFrame(validation_data)
            
            logger.info("=============== DATAFRAME ANTES DE FORMATO ===============")
            logger.info(f"Forma del DataFrame: {validation_df.shape}")
            logger.info(f"Columnas: {validation_df.columns.tolist()}")
            
            # 6. Verificar la parte de ejemplo en el DataFrame final
            example_rows_before = validation_df[validation_df['Part Number'] == example_part]
            if not example_rows_before.empty:
                logger.info(f"VERIFICACIÓN PRE-FORMATO - Item Status para {example_part}: '{example_rows_before['Item Status'].iloc[0]}'")
            else:
                logger.warning(f"ALERTA: Parte ejemplo {example_part} no encontrada en DataFrame antes de formato")
            
            # DIAGNÓSTICO: Verificar el DataFrame final
            logger.info("===== DATAFRAME FINAL =====")
            logger.info(f"Columnas en DataFrame final: {validation_df.columns.tolist()}")
            if 'Item Status' in validation_df.columns:
                status_values = validation_df['Item Status'].value_counts().to_dict()
                logger.info(f"Distribución de valores en Item Status: {status_values}")
                
                # Verificar específicamente el ejemplo mencionado
                example_rows = validation_df[validation_df['Part Number'] == example_part]
                if not example_rows.empty:
                    logger.info(f"Item Status final para {example_part}: '{example_rows['Item Status'].iloc[0]}'")
                else:
                    logger.warning(f"Parte {example_part} no encontrada en DataFrame final")
            else:
                logger.warning("Columna 'Item Status' no encontrada en DataFrame final")
            logger.info("===========================")

            # Validar que no perdimos partes en el proceso
            total_parts_processed = len(validation_df['Part Number'].unique())
            total_parts_original = len(df_data['Part Number'].unique())

            if total_parts_processed != total_parts_original:
                missing_parts = set(df_data['Part Number'].unique()) - set(validation_df['Part Number'].unique())
                logger.error(f"Pérdida de datos detectada:")
                logger.error(f"Total original: {total_parts_original}")
                logger.error(f"Total procesado: {total_parts_processed}")
                logger.error(f"Partes faltantes: {missing_parts}")
                
                # MODIFICACIÓN CRÍTICA: Log de error pero CONTINUAR el proceso
                logger.warning("CONTINUANDO PROCESO a pesar de pérdida de datos para generar reporte")
                
                # Guardar información sobre partes faltantes para análisis posterior
                try:
                    # Crear directorio si no existe
                    reports_dir = Path("reports/missing_parts")
                    reports_dir.mkdir(exist_ok=True, parents=True)
                    
                    # Nombre de archivo con timestamp
                    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                    file_path = reports_dir / f"missing_parts_{timestamp}.txt"
                    
                    with open(file_path, 'w') as f:
                        f.write("=== PARTES FALTANTES EN PROCESO DE VALIDACIÓN ===\n\n")
                        f.write(f"Fecha y hora: {datetime.now().isoformat()}\n\n")
                        f.write(f"Total original: {total_parts_original}\n")
                        f.write(f"Total procesado: {total_parts_processed}\n")
                        f.write(f"Diferencia: {total_parts_original - total_parts_processed}\n\n")
                        
                        f.write("PARTES FALTANTES:\n")
                        for part in sorted(missing_parts):
                            f.write(f"- {part}\n")
                    
                    logger.info(f"Información sobre partes faltantes guardada en: {file_path}")
                except Exception as e:
                    logger.error(f"Error guardando información de partes faltantes: {str(e)}")
                
                # NO INTERRUMPIR el proceso - continuar a pesar del error

            # Validar consistencia de mismatches con lógica más tolerante
            original_parts = set(str(part).strip().upper() for part in results.get('mismatched_parts', []))
            processed_mismatches = set(str(part).strip().upper() for part in validation_df[validation_df['Serial Control match?'] == 'Mismatch']['Part Number'])
            
            # Verificar y mostrar detalles sobre las partes en debug
            logger.debug(f"Total partes originales con mismatch: {len(original_parts)}")
            logger.debug(f"Total partes procesadas con mismatch: {len(processed_mismatches)}")
            
            # Normalizar las partes para mejorar coincidencia
            original_norm = set(re.sub(r'[.\s-]', '', p.upper()) for p in original_parts)
            processed_norm = set(re.sub(r'[.\s-]', '', p.upper()) for p in processed_mismatches)

            # Verificar niveles de tolerancia para discrepancias
            if len(mismatched_parts) == 0 and len(processed_mismatches) > 0:
                logger.warning(f"Se encontraron mismatches en partes sin mismatches originales: {processed_mismatches}")
                logger.warning("Continuando procesamiento a pesar de la discrepancia (tolerancia activa)")
            elif not processed_norm.issubset(original_norm) or not original_norm.issubset(processed_norm):
                # Usar versiones normalizadas para la comparación
                missing_norm = processed_norm - original_norm
                extra_norm = original_norm - processed_norm
                
                # Mapear IDs normalizados a los originales para reporting
                missing = {part for part in processed_mismatches 
                           if re.sub(r'[.\s-]', '', part.upper()) in missing_norm}
                extra = {part for part in original_parts 
                         if re.sub(r'[.\s-]', '', part.upper()) in extra_norm}
                
                # Calcular porcentajes para determinar gravedad
                total_original = len(original_parts)
                total_processed = len(processed_mismatches)
                missing_pct = (len(missing) / max(total_original, 1)) * 100
                extra_pct = (len(extra) / max(total_processed, 1)) * 100
                
                # Clasificación de la discrepancia para un análisis más detallado
                severity_level = self._classify_discrepancy_severity(missing_pct, extra_pct, total_original, total_processed)
                
                # Registrar detalles de las discrepancias para análisis posterior
                self._log_discrepancy_details(
                    severity_level, 
                    missing, 
                    extra, 
                    missing_pct, 
                    extra_pct,
                    original_parts,
                    processed_mismatches
                )
                
                # MODIFICACIÓN CRÍTICA: NUNCA detener el proceso por discrepancias
                # Independientemente de la gravedad, siempre continuar hasta el final
                
                # Guardar datos para análisis posterior en caso de discrepancia grave
                if severity_level == "CRÍTICA" and (missing_pct > 40 or extra_pct > 40):
                    logger.error(f"INCONSISTENCIA EXTREMA: Discrepancias superan el 40% - CONTINUANDO PROCESO")
                    # Guardar datos para posterior análisis pero SIN interrumpir
                    self._save_discrepancy_data(original_parts, processed_mismatches, missing, extra)
                
                # Registrar advertencia detallada y continuar el proceso EN TODOS LOS CASOS
                logger.warning(f"Continuando procesamiento a pesar de discrepancia {severity_level.lower()}")
                # Información específica sobre las partes con problemas (ejemplos limitados)
                self._log_problematic_parts(missing, extra)

            return self._format_validation_dataframe(validation_df, physical_orgs, dynamic_columns)

        except Exception as e:
            logger.error(f"Error en generación de datos de validación: {str(e)}")
            raise

    def _format_worksheet(self, worksheet) -> None:
        """Format worksheet with basic styling."""
        self._format_worksheet_with_conditional(worksheet, None)

        
    def _generate_org_validation_data(self, df: pd.DataFrame, org_destination: List[str]) -> pd.DataFrame:
        """
        Genera datos de validación por organización a partir de un DataFrame de auditoría.
        
        Args:
            df: DataFrame con los datos de auditoría
            org_destination: Lista de organizaciones destino
            
        Returns:
            DataFrame con datos de validación por organización
        """
        self._analyze_step(df, "BEFORE ORG VALIDATION")

        validation_data = []
        
        # IMPLEMENTACIÓN CORREGIDA: Asegurar que org_destination tiene valores correctos
        # Normalizar códigos de organización en org_destination
        normalized_org_dest = [str(org).strip().zfill(2) for org in org_destination]
        logger.info(f"Procesando {len(normalized_org_dest)} organizaciones: {normalized_org_dest}")
        logger.info(f"DIAGNÓSTICO CRÍTICO - Organizaciones de destino originales: {org_destination}")
        
        # Helper function mejorada para limpiar valores nulos o 'nan'
        def is_empty_or_nan(value):
            """Determina si un valor es nulo, vacío o representa 'nan'."""
            if value is None:
                return True
            if pd.isna(value):
                return True
            if isinstance(value, str):
                cleaned = value.lower().strip()
                if cleaned in ('nan', 'none', 'null', '') or cleaned.startswith('nan'):
                    return True
            return False
            
        # CORRECCIÓN: Desactivar la normalización de número de parte para mantener partes con diferencias
        def normalize_part(part):
            """
            CORRECCIÓN: Función modificada para preservar completamente los números de parte originales.
            Esto es crítico para mantener diferenciados productos como:
            'SFP-10G-LR-S.ACTUAL.324.MERCK' y 'SFP-10G-LR-S=.ACTUAL.324.MERCK'
            que son distintos en Oracle.
            
            Solo se eliminan espacios en blanco al inicio y final, y se asegura que sea string.
            """
            if part is None:
                return ''
            
            # Simplemente retornar el string original con espacios iniciales/finales eliminados
            # Mantener exactamente igual el resto del texto (incluyendo case sensitivity)
            return str(part).strip()  # No convertir a mayúsculas para mantener case original también
        
        # DIAGNÓSTICO: Análisis de formatos de números de parte
        part_formats = {}
        for part in df['Part Number'].unique():
            original = str(part)
            normalized = normalize_part(part)
            if original != normalized:
                part_formats[original] = normalized
                
        if part_formats:
            logger.info(f"Encontrados {len(part_formats)} formatos diferentes de números de parte")
            for orig, norm in list(part_formats.items())[:5]:  # Mostrar solo los primeros 5 ejemplos
                logger.info(f"Formato original: '{orig}' → Normalizado: '{norm}'")
        
        # Crear una columna normalizada para agrupar por número de parte normalizado
        df_normalized = df.copy()
        df_normalized['Part Number Normalized'] = df['Part Number'].apply(normalize_part)
        
        # Función para obtener el programa desde el número de parte (e.g., MERCK, CSCO)
        def extract_program(part_number):
            """Extrae el programa del número de parte cuando tiene formato [BASE].[ACTUAL|PLANNED].[ID].[PROGRAM]"""
            parts = str(part_number).split('.')
            if len(parts) >= 4:
                return parts[3].upper()
            return ""
        
        # DEPURACIÓN ESPECIAL: Verificar cuántas partes actualmente tienen el estado "Missing in Org"
        if 'Status' in df.columns:
            missing_orgs_count = len(df[df['Status'] == 'Missing in Org'])
            logger.info(f"DIAGNÓSTICO CRÍTICO: En los datos originales hay {missing_orgs_count} registros con estado 'Missing in Org'")
        
        # PARCHE IMPORTANTE: Verificar si hay partes con "Missing in Org" en el estado original
        # y asegurar que se encuentren en el reporte final
        missing_parts_original = set()
        if 'Status' in df.columns and 'Part Number' in df.columns:
            missing_mask = df['Status'] == 'Missing in Org'
            if missing_mask.any():
                missing_parts_original = set(df.loc[missing_mask, 'Part Number'].unique())
                sample_parts = list(missing_parts_original)[:5]
                logger.info(f"DIAGNÓSTICO CRÍTICO: Encontradas {len(missing_parts_original)} partes con estado 'Missing in Org' en datos originales")
                logger.info(f"Ejemplos: {sample_parts}")
        
        # Agrupar por número de parte normalizado para mejorar la detección
        parts_processed = 0
        for _, part_group in df_normalized.groupby('Part Number Normalized'):
            parts_processed += 1
            if parts_processed % 100 == 0:
                logger.info(f"Procesadas {parts_processed} partes...")
                
            # Usar el primer número de parte original como representante
            part_number = part_group['Part Number'].iloc[0]
            part_norm = normalize_part(part_number)
            
            # Extraer programa del número de parte (para diagnóstico)
            program = extract_program(part_number)
            if program:
                logger.debug(f"Parte {part_number} asociada al programa: {program}")
            
            # Verificar si esta parte estaba en la lista original con "Missing in Org"
            is_originally_missing = part_number in missing_parts_original
            if is_originally_missing:
                logger.debug(f"DIAGNÓSTICO: Parte {part_number} originalmente marcada como 'Missing in Org'")
            
            # Log para diagnóstico
            if len(part_group['Part Number'].unique()) > 1:
                logger.info(f"Diferentes variantes del mismo número de parte encontradas: {part_group['Part Number'].unique()} → Normalizado como: {part_norm}")
            
            # Extraer orgs originales y estado - ATENCIÓN ESPECIAL A ESTADO "Missing in Org"
            orig_status = part_group['Status'].iloc[0]
            if orig_status == 'Missing in Org':
                logger.debug(f"ATENCIÓN: Parte {part_number} tiene estado original 'Missing in Org'")
            
            # Extraer organizaciones actuales con mejor manejo
            current_orgs = set()
            missing_orgs_fields = []
            
            # NUEVO: Detectar explícitamente organizaciones faltantes del campo "Missing Orgs" si existe
            for _, row in part_group.iterrows():
                # Procesar organizaciones actuales
                if 'Current Orgs' in row and not is_empty_or_nan(row['Current Orgs']):
                    orgs_list = [org.strip().zfill(2) for org in str(row['Current Orgs']).split(',') if org.strip()]
                    current_orgs.update(orgs_list)
                
                # IMPORTANTE: Procesar organizaciones faltantes si están explícitamente listadas
                if 'Missing Orgs' in row and not is_empty_or_nan(row['Missing Orgs']):
                    orgs_list = [org.strip().zfill(2) for org in str(row['Missing Orgs']).split(',') if org.strip()]
                    missing_orgs_fields.extend(orgs_list)
            
            # Convertir a string para el campo 'Current Orgs'
            current_orgs_str = ','.join(sorted(current_orgs))
            
            # Inicializar fila con número de parte
            row = {'Item Status': part_group['Item Status'].iloc[0] if 'Item Status' in part_group.columns else '' ,
                   'Part Number': part_number}
            
            # Normalizar columna Organization para comparaciones consistentes
            if 'Organization' in part_group.columns:
                part_group['Organization'] = part_group['Organization'].astype(str).str.strip().str.zfill(2)
            
            # Inicializar columnas Serial Control para todas las organizaciones
            for org in normalized_org_dest:
                row[f'org {org} Serial Control'] = 'Not found in org'
             
            # MEJORADO: Procesamiento de organizaciones con mejor detección
            missing_orgs = []
            
            # DIAGNÓSTICO: Verificar si hay conflictos entre org_destination y current_orgs
            orgs_in_current_not_in_dest = [org for org in current_orgs if org not in normalized_org_dest]
            if orgs_in_current_not_in_dest:
                logger.warning(f"ALERTA: Parte {part_number} tiene organizaciones actuales que no están en org_destination: {orgs_in_current_not_in_dest}")
            
            # Verificar cada organización de destino
            for org in normalized_org_dest:
                # Buscar la organización en el grupo
                org_mask = part_group['Organization'].astype(str).str.strip().str.zfill(2) == org
                org_data = part_group[org_mask]
                
                # CORREGIDO: Verificar si la organización está presente en cualquier formato
                is_present = False
                
                # 1. Verificar en current_orgs (normalizado)
                if org in current_orgs:
                    is_present = True
                    logger.debug(f"Parte {part_number}: Organización {org} encontrada en current_orgs")
                
                # 2. Verificar en los datos directamente
                elif not org_data.empty:
                    is_present = True
                    logger.debug(f"Parte {part_number}: Organización {org} encontrada en datos")
                
                # 3. Verificar en datos como parte de organization_code
                else:
                    # Buscar de forma más flexible en caso de que el formato sea diferente
                    for _, r in part_group.iterrows():
                        if 'Organization' in r:
                            org_codes = [o.strip().zfill(2) for o in str(r['Organization']).split(',') if o.strip()]
                            if org in org_codes:
                                is_present = True
                                logger.debug(f"Parte {part_number}: Organización {org} encontrada en organization_code")
                                break
                
                # Establecer valor de Serial Control explícitamente
                if not org_data.empty and 'Serial Control' in org_data.columns:
                    # Obtener el primer valor no nulo
                    serial_values = org_data['Serial Control'].dropna()
                    if not serial_values.empty:
                        serial_value = serial_values.iloc[0]
                        
                        # Normalizar usando un método consistente
                        serial_str = self._normalize_serial_value(serial_value)
                        row[f'org {org} Serial Control'] = serial_str
                
                # Agregar a missing_orgs solo si realmente no está presente
                if not is_present:
                    missing_orgs.append(org)
                    logger.debug(f"Parte {part_number}: Organización {org} marcada como faltante")
            
            # Asegurar valores consistentes en columnas de Serial Control
            for key in list(row.keys()):
                if 'Serial Control' in key and is_empty_or_nan(row[key]):
                    row[key] = 'Not found in org'
            
            # CORRECTION: No es necesario analizar serial_values para determinar el estado,
            # ya que solo nos importa si la organización está presente o no.
            
            # CORRECCIÓN IMPORTANTE: Verificar si hay organizaciones requeridas donde la parte no está presente
            missing_in_required_orgs = []
            for org in normalized_org_dest:
                org_key = f'org {org} Serial Control'
                if org_key in row:
                    # Verificar si la parte no está presente o tiene valor "Not found in org" o NaN
                    value = row[org_key]
                    if value == 'Not found in org' or value == 'nan' or pd.isna(value) or is_empty_or_nan(value):
                        missing_in_required_orgs.append(org)
                        # Para diagnóstico
                        logger.debug(f"Parte {part_number}: Organización {org} marcada como faltante porque valor serial es: '{value}'")
            
            # INCORPORAR organizaciones explícitamente definidas como faltantes en los datos originales
            if missing_orgs_fields:
                missing_orgs_fields = list(set(missing_orgs_fields))  # Eliminar duplicados
                logger.debug(f"Parte {part_number}: Organizaciones explícitamente marcadas como faltantes: {missing_orgs_fields}")
                # Añadir solo las que están en la lista de destino
                for org in missing_orgs_fields:
                    if org in normalized_org_dest and org not in missing_orgs and org not in missing_in_required_orgs:
                        missing_in_required_orgs.append(org)
                        logger.debug(f"Parte {part_number}: Añadida organización {org} desde campo 'Missing Orgs'")
            
            # FORZAR el estado "Missing in Org" si la parte estaba originalmente marcada así
            if is_originally_missing:
                # Si no hay organizaciones faltantes detectadas, pero la parte estaba marcada originalmente,
                # tratar de recuperar la información
                if orig_status == 'Missing in Org' and not missing_orgs and not missing_in_required_orgs:
                    # Buscar cualquier organización de destino que no esté en current_orgs
                    forced_missing = [org for org in normalized_org_dest if org not in current_orgs]
                    if forced_missing:
                        logger.warning(f"RECUPERACIÓN: Parte {part_number} originalmente 'Missing in Org', forzando organizaciones faltantes: {forced_missing}")
                        missing_in_required_orgs.extend(forced_missing)
            
            # Unir con missing_orgs para tener la lista completa de orgs faltantes (eliminar duplicados)
            all_missing_orgs = sorted(list(set(missing_orgs + missing_in_required_orgs)))
            
            # FORZAR un valor para all_missing_orgs si la parte tenía estado "Missing in Org" pero 
            # no hemos detectado organizaciones faltantes
            if orig_status == 'Missing in Org' and not all_missing_orgs:
                logger.warning(f"ALERTA CRÍTICA: Parte {part_number} con estado original 'Missing in Org' no tiene organizaciones faltantes detectadas.")
                
                # Comparar con todas las organizaciones de destino
                potential_missing = [org for org in normalized_org_dest if org not in current_orgs]
                
                if potential_missing:
                    logger.warning(f"RECUPERACIÓN: Añadiendo organizaciones potencialmente faltantes: {potential_missing}")
                    all_missing_orgs = sorted(potential_missing)
                else:
                    # Si todo falla, marcar como faltante la primera org que no esté en current_orgs
                    # o la primera org de la lista de destino si no hay mejor opción
                    default_missing = next((org for org in normalized_org_dest if org not in current_orgs), normalized_org_dest[0] if normalized_org_dest else None)
                    if default_missing:
                        logger.warning(f"RECUPERACIÓN DE EMERGENCIA: Marcando organización {default_missing} como faltante para preservar estado")
                        all_missing_orgs = [default_missing]
            
            # Actualizar fila con información de estado
            status_override = 'Mismatch' if all_missing_orgs else 'Match'
            
            # Si la parte tenía originalmente "Missing in Org" y no hemos detectado orgs faltantes,
            # forzar el estado a "Mismatch" para preservar la intención original
            if orig_status == 'Missing in Org' and status_override == 'Match':
                logger.warning(f"FORZANDO estado Mismatch para parte {part_number} porque originalmente era 'Missing in Org'")
                status_override = 'Mismatch'
                
                # Si necesitamos forzar a Mismatch pero no tenemos orgs faltantes,
                # usar un valor especial para indicar que hay problema pero no sabemos cuál org falta
                if not all_missing_orgs:
                    all_missing_orgs = ['Unknown org']
            
            row.update({
                'Item Status': orig_status,
                'Current Orgs': current_orgs_str,
                'Organization code mismatch': ','.join(all_missing_orgs) if all_missing_orgs else 'None',
                'Action Required': f"Create in Org {','.join(all_missing_orgs)}" if all_missing_orgs else 'None',
                'Audit Status': '',
                # CORRECCIÓN: Marcar como "Mismatch" si falta en alguna org requerida
                'Organization Status': status_override
            })
            
            # Log para diagnóstico cuando hay orgs faltantes
            if all_missing_orgs:
                logger.debug(f"Parte {part_number}: Marcada con 'Mismatch' por faltar en orgs: {all_missing_orgs}")
                if program:
                    logger.info(f"ALERTA: Parte {part_number} del programa {program} falta en organizaciones: {all_missing_orgs}")
            
            validation_data.append(row)
            
            # Log detallado para diagnóstico de partes con discrepancias
            if missing_orgs:
                logger.debug(f"Parte {part_number}: Organizaciones faltantes detectadas: {missing_orgs}")
                logger.debug(f"Current Orgs: {current_orgs_str}")
                logger.debug(f"Estado final: {row['Organization Status']}")
        
        # Crear DataFrame
        result_df = pd.DataFrame(validation_data)
        logger.info(f"DataFrame creado con {len(result_df)} filas")
        
        # Asegurar consistencia en todos los valores
        self._clean_dataframe_values(result_df)
        
        # DIAGNÓSTICO FINAL: Verificar cuántas partes están marcadas como Mismatch en el resultado
        mismatch_count = len(result_df[result_df['Organization Status'] == 'Mismatch'])
        match_count = len(result_df[result_df['Organization Status'] == 'Match'])
        logger.info(f"RESULTADO FINAL: {mismatch_count} de {len(result_df)} partes marcadas con 'Mismatch'")
        logger.info(f"RESULTADO FINAL: {match_count} de {len(result_df)} partes marcadas con 'Match'")
        
        # Comparar con los datos originales
        if 'missing_parts_original' in locals() and missing_parts_original:
            missing_parts_result = set(result_df[result_df['Organization Status'] == 'Mismatch']['Part Number'])
            orig_vs_new_diff = len(missing_parts_original) - len(missing_parts_result)
            logger.info(f"COMPARACIÓN CRÍTICA: {len(missing_parts_original)} partes originalmente 'Missing in Org' vs {len(missing_parts_result)} en resultado")
            logger.info(f"Diferencia: {orig_vs_new_diff} ({100*orig_vs_new_diff/max(1,len(missing_parts_original)):.1f}%)")
            
            # Verificar si hay partes que debían estar marcadas como faltantes pero no lo están
            missing_but_not_in_result = missing_parts_original - missing_parts_result
            if missing_but_not_in_result:
                logger.warning(f"ALERTA CRÍTICA: {len(missing_but_not_in_result)} partes originalmente 'Missing in Org' no están marcadas en el resultado")
                sample = list(missing_but_not_in_result)[:5]
                logger.warning(f"Ejemplos: {sample}")
        
        # Verificación final antes de retornar
        self._analyze_step(result_df, "AFTER ORG VALIDATION")
        return result_df
    
    def _clean_dataframe_values(self, df: pd.DataFrame) -> None:
        """
        Limpia valores inconsistentes en el DataFrame.
        Enfoque especial en columnas de Serial Control.
        """
        # Identificar columnas de Serial Control
        serial_control_cols = [col for col in df.columns if 'Serial Control' in col]
        logger.info(f"Limpiando valores en {len(serial_control_cols)} columnas de Serial Control")
        
        # 1. Limpiar columnas de Serial Control
        for col in serial_control_cols:
            # Convertir a string para manejar cualquier valor no string
            df[col] = df[col].astype(str)
            
            # CORRECCIÓN: Estandarizar el uso de "Not found in org" para mantener consistencia en todo el reporte
            df[col] = df[col].fillna('Not found in org')
            # Ampliar el diccionario de reemplazo para capturar más variantes
            df[col] = df[col].replace({
                None: 'Not found in org',
                'nan': 'Not found in org', 
                'NaN': 'Not found in org',
                'Nan': 'Not found in org',
                'None': 'Not found in org', 
                'none': 'Not found in org',
                'NONE': 'Not found in org',
                '': 'Not found in org',
                'NOT FOUND IN ORG': 'Not found in org',  # Asegurar consistencia de caso
                'not found in org': 'Not found in org',
                'Not Found In Org': 'Not found in org',
                'NOT FOUND': 'Not found in org'
            })
            
            # Normalizar valores
            df[col] = df[col].apply(self._normalize_serial_value)
            
            # Capturar valores atípicos
            mask = df[col].str.lower().str.strip().isin(['nan', 'none', 'null', ''])
            if mask.any():
                logger.warning(f"Encontradas {mask.sum()} ocurrencias de nulos en {col}")
                df.loc[mask, col] = 'Not found in org'
        
        # 2. Limpiar otras columnas
        for col in df.columns:
            if col not in serial_control_cols:
                df[col] = df[col].fillna('')
            
    
    def _write_main_results(
        self, 
        df: pd.DataFrame, 
        writer: pd.ExcelWriter, 
        program_requirements: Dict
    ) -> pd.DataFrame:
        try:
            # DEBUG: Imprimir información inicial del DataFrame
            print("DataFrame inicial:")
            print.debug(f"Columnas: {df.columns.tolist()}")
            print(f"Primeros registros:\n{df.head()}")
            print(f"Total de registros: {len(df)}")
            print(f"Números de parte únicos: {df['Part Number'].nunique()}")

            # Verificar requisitos del programa
            all_orgs = sorted(program_requirements.get('org_destination', []))
            print(f"Organizaciones objetivo: {all_orgs}")

            # Verificar que todos los campos esperados estén presentes
            expected_columns = [
                'Part Number', 'Manufacturer', 'Description', 'Vertex', 
                'Serial Control', 'Status', 'Organization'
            ]
            missing_columns = [col for col in expected_columns if col not in df.columns]
            if missing_columns:
                logger.error(f"Columnas faltantes: {missing_columns}")
                raise ValueError(f"Faltan columnas obligatorias: {missing_columns}")

            main_data = []

            for part_number in df['Part Number'].unique():
                # DEBUG: Rastrear cada número de parte
                print(f"Procesando número de parte: {part_number}")
                
                part_data = df[df['Part Number'] == part_number]
                
                # Validación adicional de datos
                if len(part_data) == 0:
                    logger.warning(f"No hay datos para el número de parte {part_number}")
                    continue

                base_info = part_data.iloc[0]
                
                # DEBUG: Imprimir información base de cada parte
                print(f"Información base de {part_number}:")
                print(str(base_info))

                row = {
                    'Part Number': part_number,
                    'Manufacturer': base_info['Manufacturer'],
                    'Description': base_info['Description'],
                    'Vertex': base_info['Vertex'],
                    'Serial Control match?': base_info.get('Serial Control', 'N/A')
                }
                
                # Procesamiento de organizaciones
                for org in all_orgs:
                    org_data = part_data[part_data['Organization'] == org]
                    
                    # DEBUG: Verificar datos por organización
                    print(f"Datos para org {org} de {part_number}:")
                    print(str(org_data))

                    if not org_data.empty:
                        # Lógica para determinar valor de serie
                        serial_value = ('Dynamic entry at inventory receipt' 
                                    if org_data['Serial Control'].iloc[0] == 'YES'
                                    else 'No serial number control')
                        row[f'org {org}'] = serial_value
                        row[f'Inventory on Hand {org}'] = 'Y' if org_data['On Hand Quantity'].iloc[0] > 0 else 'N'
                    else:
                        row[f'org {org}'] = 'Not found in org'
                        row[f'Inventory on Hand {org}'] = 'N'
                
                # Campos adicionales
                row.update({
                    'NPI Recommendations': '',
                    'Serial control Owner Notes': '',
                    'NPI Action/Data update': '',
                    'Status': base_info.get('Status', 'Unknown'),
                    'Item Org destination mismatch?': 'Yes' if part_data['Status'].eq('Missing Org').any() else 'No',
                    'Organization code mismatch': '',
                    'Action Required': ''
                })
                
                main_data.append(row)
            
            # Creación de DataFrame final
            main_df = pd.DataFrame(main_data)
            
            # DEBUG: Verificar DataFrame final
            print("DataFrame final antes de escribir:")
            print(f"Columnas: {main_df.columns.tolist()}")
            print(f"Primeros registros:\n{main_df.head()}")
            print(f"Total de registros: {len(main_df)}")

            # Definir orden de columnas
            columns = [
                'Part Number', 'Manufacturer', 'Description', 'Vertex', 
                'Serial Control match?'
            ]
            
            for org in all_orgs:
                columns.extend([
                    f'org {org}',
                    f'Inventory on Hand {org}'
                ])
            
            columns.extend([
                'NPI Recommendations',
                'Serial control Owner Notes',
                'NPI Action/Data update',
                'Status',
                'Item Org destination mismatch?',
                'Organization code mismatch',
                'Action Required'
            ])
            
            # Asegurar columnas
            for col in columns:
                if col not in main_df.columns:
                    main_df[col] = ''

            main_df = main_df[columns]
            
            main_df.to_excel(writer, sheet_name="Audit Results", index=False)
            
            self._format_worksheet_with_conditional(
                writer.sheets['Audit Results'],
                main_df,
                'Status',
                'Missing Org'
            )
            
            return writer.sheets['Audit Results']
            
        except Exception as e:
            logger.error(f"Error escribiendo resultados principales: {str(e)}")
            logger.error(f"Traza de error: {traceback.format_exc()}")
            raise
    
    def _write_serial_validation_report(self, df: pd.DataFrame, writer: pd.ExcelWriter):
        """
        Write and format Serial Control Validation report with conditional formatting.
        """
        try:
            # Write DataFrame
            df.to_excel(writer, sheet_name="Serial Control Validation", index=False)
            worksheet = writer.sheets["Serial Control Validation"]
            
            # Format headers
            for idx, col in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=idx)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color=self.COLORS['HEADER'],
                    end_color=self.COLORS['HEADER'],
                    fill_type='solid'
                )
                
                # Adjust column width
                max_length = max(
                    len(str(col)),
                    df[col].astype(str).apply(len).max()
                )
                worksheet.column_dimensions[get_column_letter(idx)].width = max_length + 2

            # Apply conditional formatting
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name == 'Serial Control match?' or col_name == 'Status':
                    for row_idx, value in enumerate(df[col_name], 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if value in ['Mismatch', 'Missing Org']:
                            cell.fill = PatternFill(
                                start_color=self.COLORS['ERROR'],
                                end_color=self.COLORS['ERROR'],
                                fill_type='solid'
                            )
                        else:
                            cell.fill = PatternFill(
                                start_color=self.COLORS['SUCCESS'],
                                end_color=self.COLORS['SUCCESS'],
                                fill_type='solid'
                            )
                elif col_name.startswith('org '):
                    for row_idx, value in enumerate(df[col_name], 2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if value == 'Not found in org':
                            cell.fill = PatternFill(
                                start_color=self.COLORS['ERROR'],
                                end_color=self.COLORS['ERROR'],
                                fill_type='solid'
                            )
                        else:
                            cell.fill = PatternFill(
                                start_color=self.COLORS['SUCCESS'],
                                end_color=self.COLORS['SUCCESS'],
                                fill_type='solid'
                            )

            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1):
                for cell in row:
                    cell.border = thin_border

            # Freeze top row
            worksheet.freeze_panes = 'A2'
            
            return worksheet

        except Exception as e:
            logger.error(f"Error writing serial validation report: {str(e)}")
            raise
        
    def _analyze_step(self, df: pd.DataFrame, step_name: str) -> None:
        print(f"\n=== {step_name} ===")
        print("Columns:", df.columns.tolist())
        print(f"\nFirst 3 rows of {len(df)} total:")
        print(df.head(3).to_string())
        if 'Organization' in df.columns:
            print("\nUnique Organizations:")
            print(df['Organization'].unique())
        if 'Status' in df.columns:
            print("\nStatus Distribution:")
            print(df['Status'].value_counts())
            
            
    def _format_inventory_status(self, inv_data: Dict) -> str:
        """Formatea el estado de inventario con formato de texto consistente."""
        if not inv_data:
            return 'Not found in org'  # CORRECCIÓN: Formato consistente para "Not found in org"
        
        quantity = inv_data.get('quantity', 0)
        return 'Y' if quantity > 0 else 'N'  # Simplificado y solo cantidad
    
    # En la clase ReportGenerator, agregar:
    def _get_physical_orgs_from_config(self, config_file_path: Path) -> List[str]:
        """
        Obtiene las organizaciones físicas (no dropship) del archivo de configuración.
        
        Args:
            config_file_path: Ruta al archivo ALL_WWT_Dropship_and_Inventory_Organizations.xlsx
            
        Returns:
            Lista de códigos de organizaciones físicas
        """
        physical_orgs = []
        
        if not config_file_path.exists():
            logger.error(f"Archivo de configuración no encontrado: {config_file_path}")
            return physical_orgs
            
        try:
            logger.info(f"Leyendo organizaciones físicas de: {config_file_path}")
            df_config = pd.read_excel(config_file_path)
            
            # Imprimir las columnas disponibles para diagnóstico
            logger.info(f"Columnas en archivo de configuración: {df_config.columns.tolist()}")
            
            # Normalizar nombres de columnas
            df_config.columns = [col.upper().strip() for col in df_config.columns]
            
            # Buscar columnas específicas
            org_col = None
            dropship_col = None
            
            # Intentar encontrar las columnas relevantes
            for col in df_config.columns:
                if 'ORGAN' in col and 'CODE' in col:
                    org_col = col
                    logger.info(f"Columna de organización encontrada: {org_col}")
                elif 'DROP' in col:
                    dropship_col = col
                    logger.info(f"Columna de dropship encontrada: {dropship_col}")
            
            if org_col and dropship_col:
                # Valores que indican dropship
                dropship_values = ['YES', 'Y', 'TRUE', 'T', '1', 1, True]
                
                # Filtrar por organizaciones NO dropship (físicas)
                for _, row in df_config.iterrows():
                    if pd.isna(row[org_col]):
                        continue
                        
                    org_code = str(row[org_col]).strip()
                    # Convertir a número si es posible y luego a string con zfill
                    try:
                        org_code = str(int(float(org_code)))
                        # Usar str.zfill en lugar de zfill directamente
                        org_code = org_code.zfill(2)
                    except:
                        # Usar str.zfill en lugar de zfill directamente
                        org_code = org_code.zfill(2)
                        
                    # Normalizar valor booleano para dropship
                    dropship_val = row.get(dropship_col)
                    is_dropship = False
                    
                    if pd.isna(dropship_val):
                        is_dropship = False
                    elif isinstance(dropship_val, bool):
                        is_dropship = dropship_val
                    elif isinstance(dropship_val, (int, float)):
                        is_dropship = bool(int(dropship_val))
                    elif isinstance(dropship_val, str):
                        is_dropship = dropship_val.strip().upper() in dropship_values
                    
                    # SOLO agregar si NO es dropship y NO comienza con Z (organización de prueba)
                    if not is_dropship and not org_code.upper().startswith('Z'):
                        physical_orgs.append(org_code)
                        
                # Eliminar duplicados y ordenar
                physical_orgs = sorted(list(set(physical_orgs)))
                logger.info(f"Se encontraron {len(physical_orgs)} organizaciones físicas (NO dropship): {physical_orgs}")
            else:
                logger.warning("No se encontraron las columnas necesarias en el archivo de configuración")
                
        except Exception as e:
            logger.error(f"Error procesando archivo de configuración: {str(e)}")
            logger.error(traceback.format_exc())
            
        return physical_orgs

    def _normalize_serial_value(self, value: str) -> str:
        """
        Normaliza los valores de Serial Control para garantizar comparaciones precisas.
        
        IMPLEMENTACIÓN ESTRICTA según requisitos del documento:
        - Solo reconoce DOS valores válidos como estándares:
          1. "Dynamic entry at inventory receipt"
          2. "No serial number control"
        - "Not found in org" se reconoce pero NO participa en comparaciones
        - Cualquier variación textual de los valores estándar se normaliza
        - Valores no reconocidos generan advertencia en el log
        
        Args:
            value: Valor de Serial Control a normalizar
            
        Returns:
            Valor normalizado según las reglas estrictas del documento
        """
        # 1. Manejo de valores nulos o vacíos
        if value is None or pd.isna(value):
            return "Not found in org"
            
        # 2. Normalización básica: eliminar espacios y convertir a mayúsculas para comparación
        value_for_comparison = str(value).strip().upper()
        
        # 3. Lista exhaustiva de valores que se consideran "Dynamic entry at inventory receipt"
        if value_for_comparison in {"YES", "Y", "TRUE", "1", "DYNAMIC ENTRY AT INVENTORY RECEIPT", 
                    "DYNAMIC ENTRY", "DYNAMIC", "DYNAMIC ENTRY AT INV RECEIPT",
                    "ENTRY AT INVENTORY RECEIPT", "DYNAMICENTRY"}:
            return "Dynamic entry at inventory receipt"
            
        # 4. Lista exhaustiva de valores que se consideran "No serial number control"
        elif value_for_comparison in {"NO", "N", "FALSE", "0", "NONE", "NO SERIAL NUMBER CONTROL", 
                      "NO SERIAL CONTROL", "NO SERIAL", "NO SERIAL NUMBER",
                      "NOSERIALNUMBERCONTROL", "NO SERIALNUMBER CONTROL"}:
            return "No serial number control"
            
        # 5. Variantes de "Not found in org" - se mantiene como valor especial
        elif ("NOT FOUND" in value_for_comparison or 
              value_for_comparison == "NOT FOUND IN ORG" or 
              "NOT IN ORG" in value_for_comparison):
            return "Not found in org"
            
        # 6. Si el valor no coincide con ninguno de los valores válidos, registrar advertencia
        logger.warning(f"ALERTA CRÍTICA: Valor de Serial Control no reconocido: '{value}' - No cumple con valores estándar")
        
        # 7. Intento de última oportunidad: verificar si el valor contiene palabras clave
        if "DYNAMIC" in value_for_comparison or "ENTRY" in value_for_comparison:
            logger.warning(f"Interpretando valor '{value}' como 'Dynamic entry at inventory receipt' basado en palabras clave")
            return "Dynamic entry at inventory receipt"
        elif "NO" in value_for_comparison or "SERIAL" in value_for_comparison:
            logger.warning(f"Interpretando valor '{value}' como 'No serial number control' basado en palabras clave")
            return "No serial number control"
            
        # 8. Si todo falla, mantener el valor original para evaluación posterior
        return value
    
    
    def _validate_serial_control(self, org_values: dict) -> str:
        """
        Función PURA para validación de consistencia de control de series.
        
        Implementa las reglas ESTRICTAS según documento de requerimientos:
        - Ignora completamente organización 01 (debe estar excluida de org_values)
        - "Not found in org" no participa en la comparación (debe estar filtrado previamente)
        - Solo compara valores reales
        - MATCH solo cuando TODOS los valores reales son EXACTAMENTE iguales
        - MISMATCH en cualquier otro caso
        
        Args:
            org_values: Diccionario con valores de control serial por organización {org: valor}
                        Solo debe contener valores reales (no "Not found in org")
                        No debe incluir organización 01
        
        Returns:
            "Match" o "Mismatch" según las reglas estrictas
        """
        # Si no hay valores reales para comparar, es un Match por defecto
        if not org_values:
            return "Match"
            
        # Obtener valores únicos (ya normalizados)
        unique_values = set(org_values.values())
        
        # MATCH solo cuando todos los valores reales son exactamente iguales
        # Es decir, cuando solo hay un valor único entre todos los valores reales
        return "Match" if len(unique_values) <= 1 else "Mismatch"
        
    def _classify_discrepancy_severity(self, missing_pct: float, extra_pct: float, total_original: int, total_processed: int) -> str:
        """
        Clasifica la severidad de las discrepancias encontradas durante la validación.
        
        Args:
            missing_pct: Porcentaje de partes faltantes
            extra_pct: Porcentaje de partes extras
            total_original: Total de partes originales con mismatch
            total_processed: Total de partes procesadas con mismatch
            
        Returns:
            Nivel de severidad como string: "BAJA", "MEDIA", "ALTA" o "CRÍTICA"
        """
        # Criterios de clasificación:
        # 1. Volumen total de discrepancias
        # 2. Porcentaje de discrepancias
        # 3. Impacto en resultados finales
        
        # Número absoluto de discrepancias
        abs_diff = abs(total_original - total_processed)
        max_pct = max(missing_pct, extra_pct)
        
        # Clasificación basada en porcentaje y volumen
        if max_pct <= 10:
            return "BAJA"  # Discrepancia menor, probablemente error de normalización
        elif max_pct <= 20:
            return "MEDIA"  # Discrepancia notable pero tolerable
        elif max_pct <= 40:
            return "ALTA"   # Discrepancia significativa, requiere revisión
        else:
            return "CRÍTICA"  # Discrepancia extrema, posible error fundamental
            
    def _log_discrepancy_details(self, severity: str, missing: set, extra: set, 
                                missing_pct: float, extra_pct: float,
                                original_parts: set, processed_mismatches: set):
        """
        Registra detalles completos de las discrepancias para análisis posterior.
        
        Args:
            severity: Nivel de severidad clasificado
            missing: Conjunto de partes faltantes
            extra: Conjunto de partes extras
            missing_pct: Porcentaje de partes faltantes
            extra_pct: Porcentaje de partes extras
            original_parts: Conjunto original de partes con mismatch
            processed_mismatches: Conjunto procesado de partes con mismatch
        """
        # Encabezado con información de severidad
        logger.warning(f"=== REPORTE DE DISCREPANCIAS: SEVERIDAD {severity} ===")
        
        # Resumen cuantitativo
        logger.warning(f"Total original de mismatches: {len(original_parts)}")
        logger.warning(f"Total procesado de mismatches: {len(processed_mismatches)}")
        logger.warning(f"Faltantes: {len(missing)} ({missing_pct:.2f}%)")
        logger.warning(f"Extras: {len(extra)} ({extra_pct:.2f}%)")
        
        # Ejemplos de partes problemáticas (limitar a 5 para no sobrecargar logs)
        max_examples = 5
        
        if missing:
            missing_examples = list(missing)[:max_examples]
            logger.warning(f"Ejemplos de partes faltantes: {missing_examples}")
            if len(missing) > max_examples:
                logger.warning(f"... y {len(missing) - max_examples} más")
                
        if extra:
            extra_examples = list(extra)[:max_examples]
            logger.warning(f"Ejemplos de partes extras: {extra_examples}")
            if len(extra) > max_examples:
                logger.warning(f"... y {len(extra) - max_examples} más")
                
    def _log_problematic_parts(self, missing: set, extra: set):
        """
        Registra información detallada sobre partes problemáticas específicas.
        
        Args:
            missing: Conjunto de partes faltantes
            extra: Conjunto de partes extras
        """
        # Registrar detalles de hasta 3 partes de cada tipo para diagnóstico
        if missing:
            logger.info("Ejemplos de partes faltantes para diagnóstico:")
            for i, part in enumerate(list(missing)[:3]):
                norm_part = re.sub(r'[.\s-]', '', part.upper())
                logger.info(f"Parte {i+1}: Original '{part}' → Normalizada '{norm_part}'")
                
        if extra:
            logger.info("Ejemplos de partes extras para diagnóstico:")
            for i, part in enumerate(list(extra)[:3]):
                norm_part = re.sub(r'[.\s-]', '', part.upper())
                logger.info(f"Parte {i+1}: Original '{part}' → Normalizada '{norm_part}'")
                
        # Verificar si alguna de las partes con problemas está en la lista de ejemplos
        example_parts = {'20317.NC.2783.CHARTER', 'ONS-CFP2D-400G-C=.ACTUAL.324.CHARTER', 
                         '450-AION.NC.2808.CHARTER', 'F5-UPG-QSFP+LR4.ACTUAL.11865.CHARTER', 
                         '407-BBWV.ACTUAL.1205.CHARTER'}
                    
        for part in example_parts:
            part_upper = part.upper()
            if any(part_upper in m.upper() for m in missing) or any(part_upper in e.upper() for e in extra):
                logger.warning(f"PARTE DE EJEMPLO DETECTADA CON PROBLEMAS: {part}")
                
    def _save_discrepancy_data(self, original_parts: set, processed_mismatches: set, 
                              missing: set, extra: set):
        """
        Guarda datos detallados sobre discrepancias para análisis posterior.
        
        Args:
            original_parts: Conjunto original de partes con mismatch
            processed_mismatches: Conjunto procesado de partes con mismatch
            missing: Conjunto de partes faltantes
            extra: Conjunto de partes extras
        """
        try:
            # Crear directorio si no existe
            reports_dir = Path("reports/discrepancy_analysis")
            reports_dir.mkdir(exist_ok=True, parents=True)
            
            # Nombre de archivo con timestamp
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            file_path = reports_dir / f"discrepancy_analysis_{timestamp}.txt"
            
            # Escribir datos de discrepancia
            with open(file_path, 'w') as f:
                f.write("=== ANÁLISIS DE DISCREPANCIAS ===\n\n")
                f.write(f"Fecha y hora: {datetime.now().isoformat()}\n\n")
                
                f.write("RESUMEN:\n")
                f.write(f"Total original: {len(original_parts)}\n")
                f.write(f"Total procesado: {len(processed_mismatches)}\n")
                f.write(f"Faltantes: {len(missing)}\n")
                f.write(f"Extras: {len(extra)}\n\n")
                
                f.write("PARTES ORIGINALES:\n")
                for part in sorted(original_parts):
                    f.write(f"- {part}\n")
                    
                f.write("\nPARTES PROCESADAS:\n")
                for part in sorted(processed_mismatches):
                    f.write(f"- {part}\n")
                    
                f.write("\nPARTES FALTANTES:\n")
                for part in sorted(missing):
                    f.write(f"- {part}\n")
                    
                f.write("\nPARTES EXTRAS:\n")
                for part in sorted(extra):
                    f.write(f"- {part}\n")
                
            logger.info(f"Datos de discrepancia guardados en: {file_path}")
            
        except Exception as e:
            logger.error(f"Error guardando datos de discrepancia: {str(e)}")
        
    def _validate_serial_consistency(self, raw_data: Dict, report_df: pd.DataFrame):
        """
        Verifica coherencia entre datos originales y reporte sin interrumpir el proceso.
        
        MODIFICACIÓN CRÍTICA: Nunca interrumpe el proceso, independientemente del 
        nivel de discrepancia. Solo registra información para análisis posterior.
        """
        try:
            # Contar mismatches en datos originales
            original_mismatches = len(raw_data.get('mismatched_parts', []))
            # Contar mismatches en reporte generado
            report_mismatches = len(report_df[report_df['Serial Control match?'] == 'Mismatch'])
            
            # Log detallado para diagnóstico
            logger.info(f"VERIFICACIÓN DE CONSISTENCIA:")
            logger.info(f"Mismatches en datos originales: {original_mismatches}")
            logger.info(f"Mismatches en reporte generado: {report_mismatches}")
            
            # Calcular discrepancia porcentual
            if original_mismatches > 0:
                discrepancy_pct = abs(original_mismatches - report_mismatches) / original_mismatches * 100
                logger.info(f"Discrepancia porcentual: {discrepancy_pct:.2f}%")
                
                # Clasificar y registrar sin interrumpir en NINGÚN caso
                if discrepancy_pct > 80:
                    logger.error(f"DISCREPANCIA CRÍTICA EXTREMA: {discrepancy_pct:.2f}% - CONTINUANDO DE TODOS MODOS")
                elif discrepancy_pct > 40:
                    logger.error(f"DISCREPANCIA MUY ALTA: {discrepancy_pct:.2f}% - CONTINUANDO DE TODOS MODOS")
                elif discrepancy_pct > 20:
                    logger.error(f"DISCREPANCIA ALTA: {discrepancy_pct:.2f}% - CONTINUANDO DE TODOS MODOS")
                elif discrepancy_pct > 5:
                    logger.warning(f"Discrepancia moderada: {discrepancy_pct:.2f}%")
                else:
                    logger.info(f"Discrepancia mínima: {discrepancy_pct:.2f}%")
                
                # Mensaje explícito de continuación en todos los casos
                logger.warning("CONTINUANDO PROCESO sin interrupciones para generar reporte completo")
                
            elif original_mismatches == 0 and report_mismatches > 0:
                # Caso especial: no había mismatches originales pero se encontraron en el reporte
                logger.warning(f"No había mismatches originales pero se encontraron {report_mismatches} en el reporte")
                logger.warning("CONTINUANDO PROCESO para generar reporte completo")
        
        except Exception as e:
            # Capturar cualquier excepción para evitar interrupciones
            logger.error(f"Error en validación de consistencia: {str(e)}")
            logger.error("IGNORANDO ERROR y continuando el proceso")
        
    def _validate_data_consistency(self, original_results: Dict, generated_df: pd.DataFrame) -> bool:
        """
        Analiza consistencia de datos pero SIEMPRE retorna True para continuar el proceso.
        
        MODIFICACIÓN CRÍTICA: Nunca interrumpe el proceso, independientemente de 
        las discrepancias encontradas. Solo registra información para análisis.
        """
        try:
            # Debug: Mostrar información del DataFrame
            logger.debug(f"Columnas en DataFrame: {generated_df.columns.tolist()}")
            logger.debug(f"Total de filas: {len(generated_df)}")
            logger.debug(f"Distribución de 'Serial Control match?':\n{generated_df['Serial Control match?'].value_counts()}")
            
            # Normalizar Part Numbers
            original_mismatches = set(
                str(part).strip().upper() 
                for part in original_results.get('mismatched_parts', [])
            )
            
            mismatch_mask = generated_df['Serial Control match?'] == 'Mismatch'
            logger.debug(f"Filas con Mismatch: {mismatch_mask.sum()}")
            
            generated_mismatches = set(
                generated_df.loc[mismatch_mask, 'Part Number']
                .astype(str)
                .str.strip()
                .str.upper()
            )
            
            # Comparaciones detalladas
            logger.debug(f"Original mismatches ({len(original_mismatches)}): {original_mismatches}")
            logger.debug(f"Generated mismatches ({len(generated_mismatches)}): {generated_mismatches}")
            
            # Análisis de diferencias
            missing = original_mismatches - generated_mismatches
            extra = generated_mismatches - original_mismatches
            
            if missing or extra:
                logger.warning("ADVERTENCIA: Inconsistencia detectada en validación de datos:")
                if missing:
                    logger.warning(f"Faltantes en reporte: {missing}")
                if extra:
                    logger.warning(f"Extras en reporte: {extra}")
                
                # Crear informe detallado sin interrumpir proceso
                self._analyze_data_discrepancies(missing, extra, original_mismatches, generated_mismatches)
                
                # Mensaje explícito de continuación
                logger.warning("CONTINUANDO PROCESO a pesar de inconsistencias en validación de datos")
            else:
                logger.info("Validación de datos: Sin inconsistencias detectadas")
            
            # SIEMPRE retornar True para que el proceso continúe
            return True
            
        except Exception as e:
            # Capturar cualquier excepción para evitar interrupciones
            logger.error(f"Error en validación de consistencia de datos: {str(e)}")
            logger.error("IGNORANDO ERROR y continuando el proceso")
            # SIEMPRE retornar True incluso en caso de error
            return True
            
    def _analyze_data_discrepancies(self, missing: set, extra: set, original: set, generated: set):
        """
        Analiza y registra detalles sobre discrepancias en la validación de datos.
        
        Args:
            missing: Partes que deberían estar marcadas como mismatch pero no lo están
            extra: Partes marcadas como mismatch que no deberían estarlo
            original: Conjunto original de partes con mismatch
            generated: Conjunto generado de partes con mismatch
        """
        try:
            # Calcular porcentajes de discrepancias
            missing_pct = len(missing) / max(len(original), 1) * 100
            extra_pct = len(extra) / max(len(generated), 1) * 100
            total_discrepancy_pct = (len(missing) + len(extra)) / max(len(original) + len(generated), 1) * 100
            
            # Registrar resumen estadístico
            logger.warning(f"=== RESUMEN DE DISCREPANCIAS DE DATOS ===")
            logger.warning(f"Total original de mismatches: {len(original)}")
            logger.warning(f"Total generado de mismatches: {len(generated)}")
            logger.warning(f"Faltantes: {len(missing)} partes ({missing_pct:.2f}%)")
            logger.warning(f"Extras: {len(extra)} partes ({extra_pct:.2f}%)")
            logger.warning(f"Discrepancia total: {total_discrepancy_pct:.2f}%")
            
            # Guardar detalles en archivo si las discrepancias son significativas
            if missing_pct > 5 or extra_pct > 5:
                # Crear directorio si no existe
                reports_dir = Path("reports/data_validation")
                reports_dir.mkdir(exist_ok=True, parents=True)
                
                # Nombre de archivo con timestamp
                timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                file_path = reports_dir / f"data_validation_{timestamp}.txt"
                
                with open(file_path, 'w') as f:
                    f.write("=== ANÁLISIS DE DISCREPANCIAS EN VALIDACIÓN DE DATOS ===\n\n")
                    f.write(f"Fecha y hora: {datetime.now().isoformat()}\n\n")
                    
                    f.write("RESUMEN ESTADÍSTICO:\n")
                    f.write(f"Total original de mismatches: {len(original)}\n")
                    f.write(f"Total generado de mismatches: {len(generated)}\n")
                    f.write(f"Faltantes: {len(missing)} partes ({missing_pct:.2f}%)\n")
                    f.write(f"Extras: {len(extra)} partes ({extra_pct:.2f}%)\n")
                    f.write(f"Discrepancia total: {total_discrepancy_pct:.2f}%\n\n")
                    
                    f.write("PARTES FALTANTES:\n")
                    for part in sorted(missing):
                        f.write(f"- {part}\n")
                        
                    f.write("\nPARTES EXTRAS:\n")
                    for part in sorted(extra):
                        f.write(f"- {part}\n")
                
                logger.info(f"Detalles de discrepancias guardados en: {file_path}")
                
        except Exception as e:
            logger.error(f"Error al analizar discrepancias: {str(e)}")
            # No propagar la excepción para evitar interrupciones
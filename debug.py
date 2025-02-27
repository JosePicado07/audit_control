import asyncio
import sys
from pathlib import Path
import logging
import traceback
import warnings
from typing import Dict, Any
from dataclasses import dataclass
import pandas as pd
import openpyxl
from openpyxl import load_workbook

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.styles.stylesheet')

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('debug.log', mode='w', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger(__name__)

@dataclass
class ExcelMetadata:
    """Clase para almacenar metadata del archivo Excel"""
    sheet_names: list[str]
    active_sheet: str
    total_rows: int
    total_columns: int
    columns: list[str]
    sample_data: pd.DataFrame
    column_mapping: Dict[str, str]
    missing_columns: list[str]

class ExcelHandler:
    """Clase mejorada para manejar archivos Excel"""
    def __init__(self):
        # Columnas requeridas con los nombres exactos del archivo
        self.required_columns = [
            'Full Part Number',
            'Organization Code',
            'Serial Number Control',
            'Item Status',
            'Vertex Product Class',
            'Customer ID'
        ]
        
        # Mapeo exacto de columnas
        self.column_mapping = {
            'Full Part Number': 'Full Part Number',
            'Organization Code': 'Organization Code',
            'Serial Number Control': 'Serial Number Control',
            'Item Status': 'Item Status',
            'Vertex Product Class': 'Vertex Product Class',
            'Customer ID': 'Customer ID'
        }

    async def analyze_excel(self, file_path: Path) -> ExcelMetadata:
        """Analiza un archivo Excel de forma completa"""
        try:
            logger.info(f"Starting detailed Excel analysis for: {file_path}")

            # Cargar el workbook para metadata
            workbook = load_workbook(filename=file_path, read_only=True)
            sheet_names = workbook.sheetnames
            active_sheet = workbook.active.title
            
            logger.info(f"Excel structure - Sheets: {sheet_names}, Active: {active_sheet}")

            # Cargar datos con pandas para análisis
            df = pd.read_excel(file_path)
            
            # Obtener información básica
            total_rows = len(df)
            total_columns = len(df.columns)
            columns = df.columns.tolist()

            logger.info(f"Data dimensions - Rows: {total_rows}, Columns: {total_columns}")
            logger.info(f"Available columns: {columns}")

            # Analizar coincidencia de columnas
            column_mapping = {}
            missing_columns = []
            
            for required_col in self.required_columns:
                if required_col in columns:
                    column_mapping[required_col] = required_col
                else:
                    missing_columns.append(required_col)

            # Obtener muestra de datos
            sample_data = df.head()

            # Log resultados del análisis
            if missing_columns:
                logger.warning(f"Missing required columns: {missing_columns}")
            else:
                logger.info("All required columns found")

            return ExcelMetadata(
                sheet_names=sheet_names,
                active_sheet=active_sheet,
                total_rows=total_rows,
                total_columns=total_columns,
                columns=columns,
                sample_data=sample_data,
                column_mapping=column_mapping,
                missing_columns=missing_columns
            )

        except Exception as e:
            logger.error(f"Error analyzing Excel file: {str(e)}")
            logger.error(traceback.format_exc())
            raise

async def main():
    try:
        test_file = Path(r"C:\Users\picadocj\Downloads\wwt_gah_direct_scheduleable (3).xlsx")
        
        if not test_file.exists():
            logger.error(f"File not found: {test_file}")
            return

        excel_handler = ExcelHandler()
        metadata = await excel_handler.analyze_excel(test_file)

        logger.info("\n=== Excel Analysis Results ===")
        logger.info(f"File: {test_file}")
        logger.info(f"Sheets: {metadata.sheet_names}")
        logger.info(f"Active Sheet: {metadata.active_sheet}")
        logger.info(f"Total Rows: {metadata.total_rows}")
        logger.info(f"Total Columns: {metadata.total_columns}")
        logger.info(f"Available Columns: {metadata.columns}")
        
        if metadata.missing_columns:
            logger.warning(f"Missing Columns: {metadata.missing_columns}")
        else:
            logger.info("All required columns found")
        
        logger.info("\nSample Data:")
        logger.info("\n" + str(metadata.sample_data))

    except Exception as e:
        logger.error(f"Application execution failed: {str(e)}")
        logger.error(traceback.format_exc())

if __name__ == "__main__":
    if sys.platform == 'win32':
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    
    asyncio.run(main())